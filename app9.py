import streamlit as st
from google import genai
from PIL import Image, ImageDraw, ImageFont
import pandas as pd
from datetime import datetime
import os
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import time
import base64
from io import BytesIO
import urllib.request
import threading
import calendar
import re
import difflib
import math
import json

# 點圖定位元件（沒裝成功就自動退回拉桿，不影響其他功能）
try:
    from streamlit_image_coordinates import streamlit_image_coordinates as st_image_coordinates
    HAS_IMG_COORDS = True
except Exception:
    HAS_IMG_COORDS = False

# 相容修補：新版 streamlit 把 image_to_url 從 streamlit.elements.image 移走，
# 導致舊版 streamlit-drawable-canvas(0.9.3) 傳 background_image 時拿不到底圖 URL → 畫布一片空白。
# 解法：把 image_to_url 補回原位置，且「直接把底圖轉成 PNG data URI」回傳，
# 完全不依賴 streamlit 內部新簽名（之前包成 LayoutConfig 反而讓底圖產生失敗），瀏覽器一定能顯示。
def _compat_image_to_url(image, width=None, clamp=False, channels="RGB",
                         output_format="PNG", image_id="", *_a, **_kw):
    try:
        img = image
        if not isinstance(img, Image.Image):
            try:
                import numpy as _np
                img = Image.fromarray(_np.asarray(img))
            except Exception:
                return ""
        if img.mode not in ("RGB", "RGBA"):
            img = img.convert("RGB")
        # drawable-canvas 會把畫布寬度當 width 傳進來；依此縮放底圖讓它正好鋪滿畫布
        if isinstance(width, int) and width > 0 and img.width != width:
            _h = max(1, int(round(img.height * width / img.width)))
            img = img.resize((width, _h), Image.LANCZOS)
        _buf = BytesIO()
        img.save(_buf, format="PNG")
        return "data:image/png;base64," + base64.b64encode(_buf.getvalue()).decode()
    except Exception:
        return ""

# ⚠️ 一定要「強制覆蓋」：新版 streamlit 其實還留著 image_to_url（只是換了簽名），
# 若用 hasattr 判斷就不會生效，drawable-canvas 仍拿到壞掉的新版函式 → 畫布永遠空白。
# 只覆蓋 streamlit.elements.image（drawable-canvas 專用的舊位置），不動 image_utils，
# 以免影響到一般 st.image 的顯示。
try:
    import streamlit.elements.image as _st_image_mod
    _st_image_mod.image_to_url = _compat_image_to_url
except Exception:
    pass

# 畫布直接編輯元件（Canva 式：看著底圖直接拖文字＋拉角縮放＋轉圓點旋轉；上方補丁修好底圖顯示後即可用）
try:
    from streamlit_drawable_canvas import st_canvas
    HAS_CANVAS = True
except Exception:
    HAS_CANVAS = False

# ==========================================
# 0. 系統資源：自動下載高質感中文字體
# ==========================================
@st.cache_resource
def get_chinese_font(size):
    font_path = "NotoSansTC-Bold.otf"
    if not os.path.exists(font_path):
        try:
            url = "https://github.com/googlefonts/noto-cjk/raw/main/Sans/OTF/TraditionalChinese/NotoSansCJKtc-Bold.otf"
            req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req, timeout=5) as response, open(font_path, 'wb') as out_file:
                out_file.write(response.read())
        except Exception: pass
    
    try:
        return ImageFont.truetype(font_path, size)
    except:
        local_fonts = ["C:\\Windows\\Fonts\\msjh.ttc", "C:\\Windows\\Fonts\\msjh.ttf", "/System/Library/Fonts/PingFang.ttc"]
        for f in local_fonts:
            if os.path.exists(f):
                try: return ImageFont.truetype(f, size)
                except: continue
        return ImageFont.load_default()

# ==========================================
# 1. 記憶功能：雲端保險箱優先 + 本機儲存備援
# ==========================================
CONFIG_FILE = "config.txt"

def load_local_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r") as f:
            lines = f.readlines()
            if len(lines) >= 2:
                return lines[0].strip(), lines[1].strip()
    return "", ""

def save_config(api, url):
    with open(CONFIG_FILE, "w") as f:
        f.write(f"{api}\n{url}")

# ✨ 隱形防護網：加上 Try-Except 避免本機找不到 Secrets 時當機
api_key = ""
sheet_url = ""

try:
    api_key = st.secrets.get("gemini_api_key", "")
    sheet_url = st.secrets.get("google_sheet_url", "")
except Exception:
    pass # 找不到 Secrets 時安靜忽略，不跳紅字

# 如果雲端沒設定，就抓本機的備用設定
if not api_key or not sheet_url:
    local_api, local_url = load_local_config()
    api_key = api_key or local_api
    sheet_url = sheet_url or local_url

# ==========================================
# 2. BearJoy 視覺佈局與終極防護 CSS
# ==========================================
st.set_page_config(page_title="BearJoy 智能客服", page_icon="✦", layout="wide")

# 🔒 登入鎖：雲端有設定 app_password 才啟用（本機無 secrets → 自動略過，方便自己電腦用）
#    沒通過密碼前 st.stop()，整個 App（含側欄、客戶名單）都不會載入。
def _require_login():
    try:
        need = st.secrets.get("app_password", "")
    except Exception:
        need = ""
    if not need or st.session_state.get("_authed"):
        return
    st.markdown("### 🔒 BearJoy 智能客服系統")
    st.caption("請輸入密碼以進入")
    with st.form("_login_form"):
        pw = st.text_input("登入密碼", type="password", label_visibility="collapsed")
        if st.form_submit_button("登入"):
            if pw == need:
                st.session_state["_authed"] = True
                st.rerun()
            else:
                st.error("密碼錯誤，請再試一次")
    st.stop()

_require_login()

st.markdown("""
<style>
    .stApp { background-color: #FAF8F5; }
    .block-container { padding-top: 2.9rem !important; padding-bottom: 1rem !important; max-width: 1100px; }

    /* ✨ 字級層次：大標 → 中標 → 小標 → 註解小字，一眼分得出輕重 */
    /* ✨ 小標統一：所有區塊標題同一大小、同一左側細色條（與 VIP 顧客戰情室一致，前面不放 emoji 圖案） */
    h1, h2, h3, h4, h5, h6 { font-size: 16.5px !important; line-height: 1.35 !important; margin: 0.35rem 0 0.3rem 0 !important; color: #4A4238 !important; font-weight: 700 !important; letter-spacing: 0.3px !important; }
    h3, h4, h5, h6 { border-left: 3px solid #B7A98C !important; padding-left: 9px !important; }
    [data-testid="stMarkdownContainer"] p { margin-bottom: 0.35rem !important; line-height: 1.5 !important; color: #4A4238 !important; }
    /* 註解小字：說明性文字，灰、小、不搶眼 */
    [data-testid="stCaptionContainer"] { font-size: 11.5px !important; color: #A39C90 !important; line-height: 1.4 !important; }
    hr { margin: 0.7rem 0 !important; border-color: #E6E2D8 !important; }
    [data-testid="stVerticalBlock"] { gap: 0.55rem !important; }
    /* 分頁籤縮小、清爽，選中態更明顯 */
    [data-testid="stTabs"] button[role="tab"] { font-size: 14px !important; padding: 7px 14px !important; font-weight: 600 !important; }
    [data-testid="stTabs"] button[role="tab"][aria-selected="true"] { color: #4A4238 !important; font-weight: 700 !important; }
    /* 三個分頁之間多空約兩個字的距離 */
    [data-testid="stTabs"] [data-baseweb="tab-list"] { gap: 1.9rem !important; }

    [data-testid="stSidebar"] { background-color: #F0EDE5 !important; }
    /* 側欄標題（✦ BearJoy 導航）不要左側線條 */
    [data-testid="stSidebar"] h1, [data-testid="stSidebar"] h2, [data-testid="stSidebar"] h3,
    [data-testid="stSidebar"] h4, [data-testid="stSidebar"] h5, [data-testid="stSidebar"] h6 {
        border-left: none !important; padding-left: 0 !important;
    }
    .stButton>button { background-color: #798571 !important; color: white !important; border-radius: 6px !important; }
    [data-testid="stImage"] { display: flex; justify-content: center; }
    /* ✨ 區塊說明的小「?」popover 鈕：做成小圓鈕、不搶眼 */
    [data-testid="stPopover"] button {
        background: #F3F1EA !important; color: #8A8275 !important;
        border: 1px solid #DED8CC !important; border-radius: 50% !important;
        width: 30px !important; min-width: 30px !important; max-width: 30px !important;
        height: 30px !important; min-height: 30px !important;
        padding: 0 !important; font-size: 14px !important; font-weight: bold !important;
        display: inline-flex !important; align-items: center !important; justify-content: center !important;
        margin: -4px 0 6px 2px !important;
    }
    [data-testid="stPopover"] button:hover { background: #E6E2D8 !important; color: #4A4238 !important; }
    /* ✨ 功能卡片：整塊卡其底色＋外框，標題與內容同屬一個功能面板（沉睡客/優點分析/好評圖/VIP 一致） */
    div[data-testid="stVerticalBlockBorderWrapper"] {
        background: #EFEBE2 !important;
        border: 1px solid #DDD6C8 !important;
        border-radius: 14px !important;
        padding: 4px 14px 12px 14px !important;
        margin-bottom: 14px !important;
    }
    /* 卡片內的可捲動框維持白底，與卡其卡片區隔（縮圖/評價清單看得清楚） */
    div[data-testid="stVerticalBlockBorderWrapper"] div[data-testid="stVerticalBlockBorderWrapper"] {
        background: #FFFFFF !important;
    }
    /* ✨ 勾選框文字一排不換行（回購語氣／保存截圖） */
    [data-testid="stCheckbox"] label { white-space: nowrap !important; }
    [data-testid="stCheckbox"] label p { white-space: nowrap !important; font-size: 13.5px !important; margin: 0 !important; }
    /* ✨ 可複製區塊（建議回覆範本／私訊內容）：字放大、行距拉開、長行自動換行不要橫向捲 */
    div[data-testid="stCode"] pre, div[data-testid="stCode"] code,
    .stCode pre, .stCode code {
        font-size: 14.5px !important; line-height: 1.8 !important;
        white-space: pre-wrap !important; word-break: break-word !important;
    }
    div[data-testid="stCode"] pre { padding: 14px 16px !important; }
    /* ✨ 折價券文字「微調鍵」：手機不用拖曳也能移動／縮放／旋轉（拖曳把手太小很難按） */
    div[data-testid="stHorizontalBlock"]:has(.nudge-row) {
        gap: 6px !important; flex-wrap: nowrap !important; align-items: center !important;
        margin-bottom: 4px !important;
    }
    div[data-testid="stHorizontalBlock"]:has(.nudge-row) > div:is([data-testid="column"],[data-testid="stColumn"]) {
        flex: 1 1 0 !important; min-width: 0 !important; padding: 0 !important;
    }
    div[data-testid="stHorizontalBlock"]:has(.nudge-row) div.stButton > button {
        height: 50px !important; min-height: 50px !important;
        font-size: 19px !important; font-weight: bold !important;
        padding: 0 !important; border-radius: 10px !important;
        background: #A38F5A !important; white-space: nowrap !important;
    }
    div[data-testid="stHorizontalBlock"]:has(.nudge-row) div.stButton > button:hover { background: #8E7C4C !important; }

    .main-title-box {
        background: #EFEBE2;
        padding: 20px 22px; border-radius: 10px; margin: 12px 0 16px 0;
        border: 1px solid #DED8CC; box-sizing: border-box; width: 100%; overflow: visible;
        display: flex; align-items: center; justify-content: center; min-height: 70px;
    }
    .main-title-text { color: #4A4238 !important; margin: 0 !important; padding: 0 !important; font-weight: bold !important; font-size: 23px !important; letter-spacing: 1px !important; line-height: 1.5 !important; border: none !important; text-align: center !important; }
    .sub-title-text { color: #4A4238; font-weight: bold; font-size: 14px; margin: 0; }
    
    div[data-testid="stFileUploader"] { margin-bottom: -15px !important; margin-top: 5px !important; }
    div[data-testid="stExpander"] { margin-bottom: 5px !important; }
    
    /* ========================================================= */
    /* ✨ 錨點魔法 1：下載鈕與操控按鈕，絕對強制「水平同一行」 */
    /* ========================================================= */
    div[data-testid="stHorizontalBlock"]:has(.inline-row-btn):not(:has(.coupon-grid-anchor)) {
        display: flex !important;
        flex-direction: row !important;
        flex-wrap: nowrap !important;
        align-items: center !important;
        margin-top: 2px !important;
        margin-bottom: 6px !important;
        gap: 8px !important;
        width: 100% !important;
    }
    div[data-testid="stHorizontalBlock"]:has(.inline-row-btn):not(:has(.coupon-grid-anchor)) > div:is([data-testid="column"],[data-testid="stColumn"]):nth-child(1) {
        flex: 1 1 0 !important;
        width: auto !important;
        min-width: 0 !important;
        padding: 0 !important;
    }

    /* 強制鎖死後三個欄位為 40px (手機手指好按) */
    div[data-testid="stHorizontalBlock"]:has(.inline-row-btn):not(:has(.coupon-grid-anchor)) > div:is([data-testid="column"],[data-testid="stColumn"]):nth-child(n+2) {
        flex: 0 0 40px !important;
        width: 40px !important;
        min-width: 40px !important;
        max-width: 40px !important;
        padding: 0 !important;
    }
    
    /* 下載按鈕的精準對齊 */
    div[data-testid="stHorizontalBlock"]:has(.inline-row-btn):not(:has(.coupon-grid-anchor)) .stDownloadButton { margin: 0 !important; width: 100% !important;}
    div[data-testid="stHorizontalBlock"]:has(.inline-row-btn):not(:has(.coupon-grid-anchor)) .stDownloadButton button {
        height: 40px !important; min-height: 40px !important; padding: 0 10px !important;
        width: 100% !important; margin: 0 !important; font-size: 14px !important;
        display: flex !important; align-items: center !important; justify-content: center !important;
    }
    
    /* ✨ 小正方形操控按鈕：統一底色塊＋白字，尺寸一致、文字置中 */
    div[data-testid="stHorizontalBlock"]:has(.inline-row-btn):not(:has(.coupon-grid-anchor)) > div:is([data-testid="column"],[data-testid="stColumn"]):nth-child(n+2) div.stButton > button {
        background: #798571 !important;
        border: none !important;
        border-radius: 6px !important;
        color: #FFFFFF !important;
        font-size: 18px !important;
        font-weight: bold !important;
        padding: 0 !important;
        width: 40px !important; min-width: 40px !important; max-width: 40px !important;
        height: 40px !important; min-height: 40px !important; max-height: 40px !important;
        display: flex !important; justify-content: center !important; align-items: center !important;
        margin: 0 !important;
        box-shadow: none !important;
    }
    div[data-testid="stHorizontalBlock"]:has(.inline-row-btn):not(:has(.coupon-grid-anchor)) > div:is([data-testid="column"],[data-testid="stColumn"]):nth-child(n+2) div.stButton > button:hover {
        background: #687560 !important; color: #FFFFFF !important;
    }
    div[data-testid="stHorizontalBlock"]:has(.inline-row-btn):not(:has(.coupon-grid-anchor)) > div:is([data-testid="column"],[data-testid="stColumn"]):nth-child(n+2) div.stButton > button:disabled {
        opacity: 1 !important; color: #FFFFFF !important;
    }
    /* ✨ 三鍵大地色系（每個版位都一致）：↑上移＝卡其、↓下移＝大地綠、✕刪除＝大地紅，白字 */
    /* ↑ 上移：卡其 */
    div[data-testid="stHorizontalBlock"]:has(.inline-row-btn):not(:has(.coupon-grid-anchor)) > div:is([data-testid="column"],[data-testid="stColumn"]):nth-child(2) div.stButton > button {
        background: #A38F5A !important; color: #FFFFFF !important;
    }
    div[data-testid="stHorizontalBlock"]:has(.inline-row-btn):not(:has(.coupon-grid-anchor)) > div:is([data-testid="column"],[data-testid="stColumn"]):nth-child(2) div.stButton > button:hover {
        background: #8E7C4C !important;
    }
    /* ↓ 下移：大地綠 */
    div[data-testid="stHorizontalBlock"]:has(.inline-row-btn):not(:has(.coupon-grid-anchor)) > div:is([data-testid="column"],[data-testid="stColumn"]):nth-child(3) div.stButton > button {
        background: #798571 !important; color: #FFFFFF !important;
    }
    div[data-testid="stHorizontalBlock"]:has(.inline-row-btn):not(:has(.coupon-grid-anchor)) > div:is([data-testid="column"],[data-testid="stColumn"]):nth-child(3) div.stButton > button:hover {
        background: #687560 !important;
    }
    /* ✕ 刪除：大地紅 */
    div[data-testid="stHorizontalBlock"]:has(.inline-row-btn):not(:has(.coupon-grid-anchor)) > div:is([data-testid="column"],[data-testid="stColumn"]):nth-child(4) div.stButton > button {
        background: #B0746A !important; color: #FFFFFF !important;
    }
    div[data-testid="stHorizontalBlock"]:has(.inline-row-btn):not(:has(.coupon-grid-anchor)) > div:is([data-testid="column"],[data-testid="stColumn"]):nth-child(4) div.stButton > button:hover {
        background: #9C6359 !important;
    }

    /* ========================================================= */
    /* ✨ 錨點魔法 2：文字方框拉長，顏色調整紐靠右，絕對強制「水平同一行」 */
    /* ========================================================= */
    div[data-testid="stHorizontalBlock"]:has(.inline-row-txt):not(:has(.coupon-grid-anchor)) {
        display: flex !important;
        flex-direction: row !important;
        flex-wrap: nowrap !important;
        align-items: flex-end !important;
        margin-bottom: 0px !important;
        gap: 8px !important;
        width: 100% !important;
    }
    div[data-testid="stHorizontalBlock"]:has(.inline-row-txt):not(:has(.coupon-grid-anchor)) > div:is([data-testid="column"],[data-testid="stColumn"]):nth-child(1) {
        flex: 1 1 0 !important;
        width: auto !important;
        min-width: 0 !important;
        padding: 0 !important;
    }
    div[data-testid="stHorizontalBlock"]:has(.inline-row-txt):not(:has(.coupon-grid-anchor)) > div:is([data-testid="column"],[data-testid="stColumn"]):nth-child(2) {
        flex: 0 0 48px !important;
        width: 48px !important;
        min-width: 48px !important;
        max-width: 48px !important;
        padding: 0 !important;
    }

    div[data-testid="stTextArea"] label, div[data-testid="stTextInput"] label, div[data-testid="stColorPicker"] label {
        font-size: 13px !important; font-weight: bold !important; color: #798571 !important;
        margin-bottom: 4px !important; padding: 0 !important; white-space: nowrap !important;
        height: 20px !important; line-height: 20px !important;
    }
    /* 壓印文字框（單行）高度 42px */
    div[data-testid="stTextInput"] input { height: 42px !important; padding: 6px 10px !important; }
    div[data-testid="stTextArea"] textarea { min-height: 42px !important; height: 42px !important; padding: 6px 10px !important; }
    /* 顏色色塊＝正方形，高度與左邊壓印文字框一致(42px)、底部對齊同一排 */
    div[data-testid="stColorPicker"] { margin-top: 0px !important; display: flex; flex-direction: column; justify-content: flex-end; align-items: flex-start;}
    div[data-testid="stColorPicker"] div[role="button"] { width: 42px !important; height: 42px !important; min-width: 42px !important; padding: 0 !important; border-radius: 6px !important; }

    /* ========================================================= */
    /* 拉桿(BAR)對齊優化 */
    /* ========================================================= */
    div[data-testid="stSlider"] {
        display: flex !important; flex-direction: row !important; flex-wrap: nowrap !important;
        align-items: center !important; margin-bottom: -5px !important;
    }
    div[data-testid="stSlider"] > label {
        flex: 0 0 auto !important; width: 60px !important; margin-bottom: 0px !important;
        margin-right: 5px !important; font-size: 13px !important; font-weight: bold !important;
        color: #798571 !important; white-space: nowrap !important;
    }
    div[data-testid="stSlider"] > div { flex: 1 1 auto !important; }

    /* ========================================================= */
    /* ✨ 手機優化：折價券改單欄全寬，按鈕排才不會被擠到跑版 */
    /* ========================================================= */
    /* 🔧 隱形排版標記（keep-row/ratio-row/trio-btn/main-stack）所在的 markdown 容器整個收掉，
       不可佔任何高度，否則會把該欄內容往下推、害同排另一欄顯得「太高」沒對齊 */
    [data-testid="stElementContainer"]:has(> div[data-testid="stMarkdown"] span.keep-row),
    [data-testid="stElementContainer"]:has(> div[data-testid="stMarkdown"] span.ratio-row),
    [data-testid="stElementContainer"]:has(> div[data-testid="stMarkdown"] span.trio-btn),
    [data-testid="stElementContainer"]:has(> div[data-testid="stMarkdown"] span.cck-row),
    [data-testid="stElementContainer"]:has(> div[data-testid="stMarkdown"] span.main-stack) {
        display: none !important;
        height: 0 !important;
        min-height: 0 !important;
        margin: 0 !important;
        padding: 0 !important;
    }

    /* 回購語氣／保存截圖：兩顆貼齊左側、各取內容寬、隔約一個字距、垂直置中；保存截圖再往上微調對齊 */
    div[data-testid="stHorizontalBlock"]:has(.cck-row):not(:has(.main-stack)) {
        flex-wrap: nowrap !important;
        gap: 0.85rem !important;
        align-items: center !important;
        justify-content: flex-start !important;
    }
    div[data-testid="stHorizontalBlock"]:has(.cck-row):not(:has(.main-stack)) > div:is([data-testid="column"],[data-testid="stColumn"]) {
        min-width: 0 !important;
        flex: 0 0 auto !important;
        width: auto !important;
    }
    div[data-testid="stHorizontalBlock"]:has(.cck-row):not(:has(.main-stack)) > div:is([data-testid="column"],[data-testid="stColumn"]):nth-child(2) {
        margin-top: -3px !important;
    }

    @media (max-width: 820px) {
        /* 折價券左右兩欄 → 改成上下單欄，每張券吃滿整個螢幕寬 */
        div[data-testid="stHorizontalBlock"]:has(.coupon-grid-anchor) {
            flex-direction: column !important;
            gap: 0 !important;
        }
        div[data-testid="stHorizontalBlock"]:has(.coupon-grid-anchor) > div:is([data-testid="column"],[data-testid="stColumn"]) {
            flex: 1 1 100% !important;
            width: 100% !important;
            min-width: 100% !important;
            max-width: 100% !important;
        }
        /* 三顆大地綠按鈕（重新整理／產生好評圖／打包原圖）→ 手機維持三顆同一排，文字不換行（縮小字級塞進一行） */
        div[data-testid="stHorizontalBlock"]:has(.trio-btn) {
            flex-wrap: nowrap !important;
            gap: 0.3rem !important;
        }
        div[data-testid="stHorizontalBlock"]:has(.trio-btn) > div:is([data-testid="column"],[data-testid="stColumn"]) {
            min-width: 0 !important;
            flex: 1 1 0 !important;
        }
        div[data-testid="stHorizontalBlock"]:has(.trio-btn) div.stButton > button {
            white-space: nowrap !important;
            font-size: 12px !important;
            padding-left: 2px !important;
            padding-right: 2px !important;
        }
        div[data-testid="stHorizontalBlock"]:has(.trio-btn) div.stButton > button p {
            white-space: nowrap !important;
        }
        /* 第一步驟上傳區 ＋ 第二步驟結果區 → 手機上下單欄，框吃滿整個螢幕寬（恢復原本寬度） */
        div[data-testid="stHorizontalBlock"]:has(.main-stack) {
            flex-direction: column !important;
            gap: 0 !important;
        }
        div[data-testid="stHorizontalBlock"]:has(.main-stack) > div:is([data-testid="column"],[data-testid="stColumn"]) {
            flex: 1 1 100% !important;
            width: 100% !important;
            min-width: 100% !important;
            max-width: 100% !important;
        }
        /* 大小/旋轉、左右/上下 這幾組拉桿在手機上也改成單欄，比較好拖 */
        div[data-testid="stHorizontalBlock"]:has(.slider-pair-anchor):not(:has(.coupon-grid-anchor)) {
            flex-direction: column !important;
            gap: 0 !important;
        }
        div[data-testid="stHorizontalBlock"]:has(.slider-pair-anchor):not(:has(.coupon-grid-anchor)) > div:is([data-testid="column"],[data-testid="stColumn"]) {
            flex: 1 1 100% !important;
            width: 100% !important;
            min-width: 100% !important;
            max-width: 100% !important;
        }
        /* ✨ 手機版：只有「標記為 keep-row」的列維持並排，其餘欄位各自一排（自然堆疊）；
           排除外層 main-stack／coupon-grid 容器（那兩層要上下堆疊，不可被這條覆寫） */
        div[data-testid="stHorizontalBlock"]:has(.keep-row):not(:has(.main-stack)):not(:has(.coupon-grid-anchor)) {
            flex-wrap: nowrap !important;
            gap: 0.4rem !important;
            align-items: center !important;
        }
        div[data-testid="stHorizontalBlock"]:has(.keep-row):not(:has(.main-stack)):not(:has(.coupon-grid-anchor)) > div:is([data-testid="column"],[data-testid="stColumn"]) {
            min-width: 0 !important;
            flex: 1 1 0 !important;
        }
        /* ✨ 手機版：標記 ratio-row 的列維持並排，但保留各欄原本的寬度比例（如版型較寬、取幾筆較窄） */
        div[data-testid="stHorizontalBlock"]:has(.ratio-row):not(:has(.main-stack)):not(:has(.coupon-grid-anchor)) {
            flex-wrap: nowrap !important;
            gap: 0.4rem !important;
            align-items: center !important;
        }
        div[data-testid="stHorizontalBlock"]:has(.ratio-row):not(:has(.main-stack)):not(:has(.coupon-grid-anchor)) > div:is([data-testid="column"],[data-testid="stColumn"]) {
            min-width: 0 !important;
        }
        /* ✨ 手機版：分頁標籤縮小間距、字略小，三個分頁一頁就看完整 */
        [data-testid="stTabs"] button[role="tab"] { padding: 6px 6px !important; font-size: 12.5px !important; }
        /* 手機版分頁間距：多空約兩個字（太擠不好點，且使用者要求拉開） */
        [data-testid="stTabs"] [data-baseweb="tab-list"] { gap: 1.4rem !important; }
    }

    /* ========================================================= */
    /* ✨ 新增：按鈕統一大小 + 折價券版位色塊區隔 */
    /* ========================================================= */
    /* ✨ 按鈕統一規格：底色塊（綠）、白字、固定高度、文字置中（折價券小方塊鍵有自己的規則，不受影響） */
    .stButton > button, .stDownloadButton > button {
        background-color: #798571 !important; color: #FFFFFF !important;
        border: none !important; border-radius: 6px !important;
        height: 42px !important; min-height: 42px !important;
        /* 按鈕填滿所在欄位（在欄位排版裡才會整齊對齊）；不另設固定寬避免跑版 */
        width: 100% !important; max-width: 100% !important;
        font-size: 14px !important; font-weight: 600 !important;
        padding: 0 14px !important;
        display: inline-flex !important; align-items: center !important; justify-content: center !important;
        text-align: center !important; line-height: 1.2 !important;
        box-shadow: none !important;
    }
    .stButton > button:hover, .stDownloadButton > button:hover {
        background-color: #687560 !important; color: #FFFFFF !important;
    }
    .stButton > button:disabled {
        background-color: #C9C4B8 !important; color: #FFFFFF !important;
    }
    /* ✨ 按鈕內所有文字一律白色、置中（含 Streamlit 把文字包進 <p>/<span> 的情況） */
    .stButton > button *, .stDownloadButton > button * { color: #FFFFFF !important; }
    .stButton > button p, .stDownloadButton > button p,
    .stButton > button div, .stDownloadButton > button div {
        margin: 0 !important; text-align: center !important;
        width: 100% !important; display: flex !important;
        align-items: center !important; justify-content: center !important;
    }
    /* 折價券那一排的下載鍵維持原欄寬，不被上面的限制縮短 */
    div[data-testid="stHorizontalBlock"]:has(.inline-row-btn) .stDownloadButton > button {
        max-width: none !important;
    }
    /* ✨ 折價券：每個版位包成米色卡片，清楚分辨哪些按鈕屬於哪個版位 */
    div:is([data-testid="column"],[data-testid="stColumn"]):has(.coupon-grid-anchor) {
        background: #F5F3EC !important;
        border: 1px solid #E6E2D8 !important;
        border-radius: 12px !important;
        padding: 14px !important;
        overflow: visible !important;
    }
    /* ✨ 折疊區（expander）：邊框完整呈現、不被卡片邊緣截掉（含「✏️ 想加日期/文字」） */
    div[data-testid="stExpander"] { overflow: visible !important; margin-top: 8px !important; }
    div[data-testid="stExpander"] > details {
        border: 1px solid #E0DBD0 !important; border-radius: 8px !important;
        background: #FFFFFF !important; overflow: visible !important;
    }
    div[data-testid="stExpander"] > details > summary { border-radius: 8px !important; }
    /* ✨ 側邊欄折疊標題（💡 開啟太慢…）：字加大、一排不換行、上下＋左右都置中 */
    [data-testid="stSidebar"] div[data-testid="stExpander"] > details > summary {
        display: flex !important; align-items: center !important; justify-content: center !important;
        min-height: 46px !important; padding-top: 0 !important; padding-bottom: 0 !important;
        position: relative !important;
    }
    /* 展開箭頭改絕對定位（不佔 flex 空間），標題文字才能在整個方框內真正左右置中 */
    [data-testid="stSidebar"] div[data-testid="stExpander"] > details > summary svg {
        position: absolute !important; left: 12px !important; top: 50% !important;
        transform: translateY(-50%) !important;
    }
    /* 標題文字框絕對定位、鋪滿整個方框後置中：徹底不受左側箭頭佔位影響，真正上下左右置中 */
    [data-testid="stSidebar"] div[data-testid="stExpander"] summary [data-testid="stMarkdownContainer"] {
        position: absolute !important; left: 0 !important; right: 0 !important;
        top: 0 !important; bottom: 0 !important;
        display: flex !important; align-items: center !important; justify-content: center !important;
        text-align: center !important; margin: 0 !important; pointer-events: none !important;
    }
    [data-testid="stSidebar"] div[data-testid="stExpander"] summary p {
        font-size: 15px !important; font-weight: bold !important; white-space: nowrap !important;
        line-height: 1.2 !important; margin: 0 !important; text-align: center !important;
        color: #4A4238 !important;
        display: flex !important; align-items: center !important; justify-content: center !important;
    }
    /* ✨ 側邊欄折疊內文：整區（含粗體標題、條列數字 1/2/3）統一同一字級、同樣行距，不雜亂 */
    [data-testid="stSidebar"] div[data-testid="stExpander"] [data-testid="stMarkdownContainer"] p,
    [data-testid="stSidebar"] div[data-testid="stExpander"] [data-testid="stMarkdownContainer"] li,
    [data-testid="stSidebar"] div[data-testid="stExpander"] [data-testid="stMarkdownContainer"] ol,
    [data-testid="stSidebar"] div[data-testid="stExpander"] [data-testid="stMarkdownContainer"] ul,
    [data-testid="stSidebar"] div[data-testid="stExpander"] [data-testid="stMarkdownContainer"] strong {
        font-size: 11.5px !important; line-height: 1.5 !important; color: #4A4238 !important;
    }
    [data-testid="stSidebar"] div[data-testid="stExpander"] [data-testid="stMarkdownContainer"] p {
        margin-bottom: 0.45rem !important;
    }
    [data-testid="stSidebar"] div[data-testid="stExpander"] [data-testid="stMarkdownContainer"] ol {
        margin: 0 0 0.45rem 0 !important; padding-left: 1.1rem !important;
    }
</style>
""", unsafe_allow_html=True)

# ==========================================
# 3. 雲端引擎與【無損高畫質儲存技術】
# ==========================================
@st.cache_resource(show_spinner=False)
def _open_sheet_cached(url):
    """✨ 速度優化：成功的連線會被快取,之後每次互動都直接重用,不再重新登入金鑰、
    重新打開試算表(這是操作卡頓的主因)。失敗時丟出例外→不快取,下次互動會自動重試。"""
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds = None
    try:
        if "type" in st.secrets:
            creds = ServiceAccountCredentials.from_json_keyfile_dict(dict(st.secrets), scope)
    except Exception:
        pass
    if not creds:
        key_path = os.path.join(os.path.dirname(__file__), "google_key.json")
        if os.path.exists(key_path):
            creds = ServiceAccountCredentials.from_json_keyfile_name(key_path, scope)
    if not creds:
        raise RuntimeError("找不到金鑰")
    gc = gspread.authorize(creds)
    return gc.open_by_url(url)

def connect_google_sheets(url):
    """對外介面不變,維持回傳 (doc, 訊息);底層改用上方快取連線。"""
    try:
        return _open_sheet_cached(url), "成功"
    except Exception as e:
        return None, str(e)

def get_or_create_ws(doc, title):
    try: return doc.worksheet(title)
    except: return doc.add_worksheet(title=title, rows="100", cols="10")

def img_to_base64_chunks(img):
    # 折價券專用：用「無損 PNG」存，文字/邊緣完全不糊（JPEG 即使 q100 仍會在文字邊緣產生鬼影）
    if img.mode in ("RGBA", "P"): img = img.convert("RGB")
    # 解析度上限拉高到 2600，確保清晰（thumbnail 只會縮小、不會放大，原圖較小則維持原樣）
    img.thumbnail((2600, 2600), Image.Resampling.LANCZOS)
    buffered = BytesIO()
    img.save(buffered, format="PNG", optimize=True)  # 無損；optimize 在不損畫質下縮小體積
    b64_str = base64.b64encode(buffered.getvalue()).decode()
    return [b64_str[i:i+45000] for i in range(0, len(b64_str), 45000)]

def base64_chunks_to_img(chunks):
    b64_str = "".join(chunks)
    return Image.open(BytesIO(base64.b64decode(b64_str)))

def img_to_chunks_compact(img, maxpx=2000, quality=95):
    """素材用：base64 切塊（高畫質，版型A拼接清晰；仍壓縮避免雲端肥大）。"""
    if img.mode in ("RGBA", "P"):
        img = img.convert("RGB")
    img.thumbnail((maxpx, maxpx), Image.Resampling.LANCZOS)
    buffered = BytesIO()
    img.save(buffered, format="JPEG", quality=quality, subsampling=0)  # subsampling=0：文字邊緣更銳利
    b64 = base64.b64encode(buffered.getvalue()).decode()
    return [b64[i:i + 45000] for i in range(0, len(b64), 45000)]

def _save_kv(ws, key, value):
    """在工作表第一欄找 key：有就更新第二欄，沒有就新增一列（即時讀取，安全）。"""
    col = ws.col_values(1)
    if key in col:
        ws.update_cell(col.index(key) + 1, 2, value)
    else:
        ws.append_row([key, value])

def write_df_to_sheet(doc, title, df):
    """把整份 DataFrame 覆蓋寫入指定 Google Sheet 工作表（沒有就新建），回傳該工作表網址。
    讓沉睡客名單、顧客優點分析也能像 VIP名單一樣，直接在雲端 Google Sheet 的新分頁查看／下載 Excel。"""
    ws = get_or_create_ws(doc, title)
    ws.clear()
    rows = [list(map(str, df.columns))] + df.astype(str).values.tolist()
    ws.append_rows(rows, value_input_option="RAW")
    # 📐 對齊：日期/時間/天數/次數欄靠右，其餘文字欄靠左；皆靠上＋自動換行展開
    try:
        n = len(rows)
        for i, name in enumerate(df.columns):
            col = chr(ord('A') + i)
            ws.format(f"{col}1:{col}{n}", {
                "horizontalAlignment": "RIGHT" if _col_align_right(name) else "LEFT",
                "verticalAlignment": "TOP", "wrapStrategy": "WRAP"})
    except Exception:
        pass
    try:
        return f"{doc.url}#gid={ws.id}"
    except Exception:
        return None

def _col_align_right(name):
    """欄位名稱含 日期/時間/天數/次數/數量/金額/互動 → 靠右（數字、日期慣例）；其餘靠左。"""
    return any(k in str(name) for k in ["日期", "時間", "天數", "次數", "數量", "金額", "互動"])

def _strip_md(text):
    """移除 Markdown 標記（**粗體**、*斜體*、# 標題、` 程式碼、條列符號），給 Excel／試算表用純文字。"""
    s = str(text)
    s = s.replace("**", "").replace("__", "")
    s = re.sub(r'(?<!\*)\*(?!\*)', '', s)
    s = re.sub(r'^#{1,6}\s*', '', s, flags=re.M)
    s = s.replace("`", "")
    s = re.sub(r'^\s*[-•*]\s+', '', s, flags=re.M)
    return s.strip()

# ==========================================
# 📚 客服問題分類庫：可分類、編輯、搜尋，並用 API 生成可直接複製的建議回覆範本
#    資料存在雲端 Google Sheet「客服問題庫」分頁，手機／電腦都能即時查看與編輯。
# ==========================================
QA_SHEET = "客服問題庫"
QA_COLS = ["ID", "分類", "問題標題", "客戶問題範例", "建議回覆範本", "關鍵字", "更新時間"]


def qa_reply_height(txt, base=420, per_line=26, cap=1400):
    """回覆框高度：依內容行數自動加長，盡量一眼看到完整內容好編輯（最高 cap 之後才出現捲軸）。
    長行也算進去（一行約 34 個字換行一次）。"""
    t = str(txt or "")
    lines = 0
    for ln in t.split("\n"):
        lines += max(1, (len(ln) // 34) + 1)
    return int(max(base, min(cap, 120 + lines * per_line)))
# 內建預設分類（第一次用就有東西可選）；之後可在介面「管理分類」自行增刪改，改完存雲端。
QA_DEFAULT_CATS = ["物流出貨", "退換貨", "商品規格", "付款發票", "折價券優惠",
                   "售後保固", "訂單修改", "客訴負評", "其他"]
QA_CATS_KEY = "qa_cats"

def qa_cats_load(doc):
    """讀分類清單（存在「系統設定」的 qa_cats，用｜分隔）；沒設定過就回傳預設清單。
    只讀第一欄＋單一格，不會把系統設定裡折價券的 base64 大資料整包抓下來。"""
    try:
        ws = get_or_create_ws(doc, "系統設定")
        col = ws.col_values(1)
        if QA_CATS_KEY in col:
            v = ws.cell(col.index(QA_CATS_KEY) + 1, 2).value or ""
            cats = [c.strip() for c in v.split("｜") if c.strip()]
            if cats:
                return cats, ws
        return list(QA_DEFAULT_CATS), ws
    except Exception:
        return list(QA_DEFAULT_CATS), None

def qa_cats_save(ws, cats):
    """存回分類清單（去重、保留順序），回傳整理後的清單。"""
    seen, out = set(), []
    for c in cats:
        c = (c or "").strip()
        if c and c not in seen:
            seen.add(c)
            out.append(c)
    _save_kv(ws, QA_CATS_KEY, "｜".join(out))
    return out

def qa_load(doc):
    """讀取客服問題庫；空表自動補表頭。回傳 (ws, list[dict])，每筆含 _row 實際列號供更新/刪除。"""
    ws = get_or_create_ws(doc, QA_SHEET)
    values = ws.get_all_values()
    # ⚠️ gspread 新建空白表的 get_all_values() 會回傳 [[]]（非 []），第一列也可能全空 →
    #    都視為「還沒有表頭」，補上表頭再回傳空清單，避免第一筆資料被塞到第 1 列造成錯位。
    if not values or not any((c or "").strip() for c in values[0]):
        ws.append_row(QA_COLS, value_input_option="RAW")
        return ws, []
    header = values[0]
    rows = []
    for i, r in enumerate(values[1:], start=2):
        if not any((c or "").strip() for c in r):  # 跳過全空列
            continue
        d = {col: (r[j] if j < len(r) else "") for j, col in enumerate(header)}
        rec = {c: d.get(c, "") for c in QA_COLS}
        rec["_row"] = i
        rows.append(rec)
    return ws, rows

def qa_add(ws, rec):
    ws.append_row([rec.get(c, "") for c in QA_COLS], value_input_option="RAW")

def qa_update(ws, row, rec):
    # gspread 6.x：values 為第一參數，這裡一律用關鍵字避免版本差異
    ws.update(values=[[rec.get(c, "") for c in QA_COLS]], range_name=f"A{row}:G{row}",
              value_input_option="RAW")

def qa_delete(ws, row):
    ws.delete_rows(row)

# 🤖 Gemini 模型清單與計價（每 100 萬 tokens 美金 in/out；free=免費額度可用，預設選免費）
#    報價依 Google 官方（2026/06）：3 Flash、3.1 Flash-Lite、2.5 Flash 系列有免費額度（額度縮減）；
#    Pro 系列 2026-04 起取消免費額度、純付費。未來新模型在此增修即可。
#    ids：同一款的候選 API 代號（preview→正式可能改名），呼叫時依序嘗試，避免改版就失效。
GEMINI_PRICES = {
    "gemini-2.5-flash":      {"label": "2.5 Flash ｜免費額度・預設穩定", "in": 0.30, "out": 2.50, "free": True,
                              "ids": ["gemini-2.5-flash"]},
    "gemini-2.5-flash-lite": {"label": "2.5 Flash-Lite ｜免費額度・便宜快", "in": 0.10, "out": 0.40, "free": True,
                              "ids": ["gemini-2.5-flash-lite"]},
    "gemini-3-flash":        {"label": "3 Flash ｜最新（需帳號已開放）", "in": 0.50, "out": 3.00, "free": True,
                              "ids": ["gemini-3-flash", "gemini-3-flash-preview"]},
    "gemini-3.1-flash-lite": {"label": "3.1 Flash-Lite ｜最新最省（需帳號已開放）", "in": 0.25, "out": 1.50, "free": True,
                              "ids": ["gemini-3.1-flash-lite", "gemini-3.1-flash-lite-preview"]},
    "gemini-2.5-pro":        {"label": "2.5 Pro ｜付費・最聰明", "in": 1.25, "out": 10.0, "free": False,
                              "ids": ["gemini-2.5-pro"]},
}
GEMINI_DEFAULT = "gemini-2.5-flash"  # 預設用「確定可用」的免費款；3 系列保留可手動選
GEMINI_SAFE_FALLBACK = "gemini-2.5-flash"  # 選的模型整個打不通時的最後保險（已知穩定免費）

def _pil_from_upload(uploaded):
    """把 Streamlit file_uploader 的上傳檔轉成 PIL Image（給多模態 API 讀截圖）；失敗回 None。"""
    if uploaded is None:
        return None
    try:
        uploaded.seek(0)
        return Image.open(uploaded).convert("RGB")
    except Exception:
        return None

def gemini_call_costed(api_key, contents, model):
    """呼叫指定 Gemini 模型，回傳 (text, usage)。
    usage = {model, in, out, cost(美金估算), free}。可重複用在任何需要 API 的功能。
    contents 可為 [文字]，或 [文字, PIL圖片]（多模態，讀截圖）。失敗回 (None, usage_or_None)。"""
    try:
        client = genai.Client(api_key=api_key)
    except Exception:
        return None, None
    # 先試指定模型的候選代號；整款都打不通時，最後退回已知穩定的免費款，確保仍生成得出來。
    order = [model] + ([GEMINI_SAFE_FALLBACK] if model != GEMINI_SAFE_FALLBACK else [])
    resp, used = None, model
    first = True
    for mkey in order:
        price = GEMINI_PRICES.get(mkey, {"in": 0.0, "out": 0.0, "free": True, "ids": [mkey]})
        for mid in price.get("ids", [mkey]):   # 候選代號（preview/正式改名也不會壞）
            for attempt in range(2):            # 每代號重試 1 次，免費版偶發流量限制時退避再試
                try:
                    if not first:
                        time.sleep(2 + attempt * 2)
                    first = False
                    resp = client.models.generate_content(model=mid, contents=contents)
                    used = mkey
                    break
                except Exception:
                    resp = None
                    continue
            if resp is not None:
                break
        if resp is not None:
            break
    if resp is None:
        return None, None
    price = GEMINI_PRICES.get(used, {"in": 0.0, "out": 0.0, "free": True})
    um = getattr(resp, "usage_metadata", None)
    pin = getattr(um, "prompt_token_count", 0) or 0
    pout = getattr(um, "candidates_token_count", 0) or 0
    cost = pin / 1_000_000 * price["in"] + pout / 1_000_000 * price["out"]
    usage = {"model": used, "in": pin, "out": pout, "cost": cost,
             "free": price.get("free", True), "fellback": used != model}
    try:
        return resp.text, usage
    except Exception:
        return None, usage

def qa_ai_suggest(api_key, category, question, model=GEMINI_DEFAULT, image=None):
    """依客戶問題（文字＋可選截圖）＋分類，用 Gemini 生成 BearJoy Sharon 語氣、可直接複製的建議回覆。
    回傳 (回覆文字 or None, usage)。"""
    q = (question or "").strip()
    if image is not None and not q:
        qline = "【客戶問題】請先辨識下方附上的截圖中，客人提出的問題與需求，再回覆。"
    elif image is not None and q:
        qline = f"【客戶問題】{q}\n（另附客人對話截圖，請一併參考圖中細節）"
    else:
        qline = f"【客戶問題】{q}"
    prompt = f"""你是蝦皮賣場 BearJoy 的客服主管 Sharon。請針對下方客戶問題，寫一則可以直接複製貼給客人的回覆。
【語氣】溫暖、專業、貼心，適度使用 Emoji，分段換行、版面清爽。
【長度】約 3～6 行，務必精簡，不要長篇大論。
【結尾】最後一行署名：—— BearJoy Sharon
【問題分類】{category or '一般客服'}
{qline}

請「只」輸出回覆內容本身，不要加任何說明、標題或引號。"""
    contents = [prompt, image] if image is not None else [prompt]
    text, usage = gemini_call_costed(api_key, contents, model)
    if not text:
        return None, usage
    return _strip_md(text).strip(), usage

def qa_ai_refine(api_key, base_reply, instruction, category="", question="",
                 model=GEMINI_DEFAULT, image=None):
    """把「現有的回覆範本」依你的加強指示改寫成一則完整、可直接複製貼給客人的回覆。
    例：想加強「多強調七天鑑賞期」「語氣再親切一點」「補一句提醒保留外箱」。
    沒有現成回覆時就等同直接生成。回傳 (新回覆 or None, usage)。"""
    base = (base_reply or "").strip()
    want = (instruction or "").strip()
    if not want:
        return None, None
    prompt = f"""你是蝦皮賣場 BearJoy 的客服主管 Sharon。請依照我的要求，改寫下面這則客服回覆。

【語氣】溫暖、專業、貼心，適度使用 Emoji，分段換行、版面清爽。
【長度】約 3～8 行，精簡好讀，不要長篇大論。
【結尾】最後一行署名：—— BearJoy Sharon

【重要規則】
1. 保留原回覆裡已經正確的資訊與承諾，不要刪掉、也不要改動條件（天數、金額、運費由誰負擔等）。
2. 把我要加強的內容**自然融入**整則回覆，不要只是加在最後一句、也不要條列成說明。
3. 不可以自己新增原本沒有的承諾（例如額外折扣、免運、送贈品）。
4. 輸出「完整的一則回覆」，是可以直接整段複製貼給客人的狀態。

【問題分類】{category or '一般客服'}
【客戶問題】{question.strip() if question else '（未填，請依原回覆內容判斷）'}

【原本的回覆】
{base if base else '（目前還沒有回覆，請直接依我的要求寫一則新的）'}

【我要加強／調整的地方】
{want}

請「只」輸出改寫後的完整回覆內容本身，不要加任何說明、標題或引號。"""
    contents = [prompt, image] if image is not None else [prompt]
    text, usage = gemini_call_costed(api_key, contents, model)
    if not text:
        return None, usage
    return _strip_md(text).strip(), usage

def _json_list_from_text(text):
    """從 AI 回覆中取出 JSON 陣列（容忍 ```json 圍欄、前後多餘說明）。失敗回 []。"""
    if not text:
        return []
    s = str(text).strip()
    s = re.sub(r"^```[a-zA-Z]*\s*", "", s)
    s = re.sub(r"\s*```$", "", s).strip()
    if not s.startswith("["):
        i, j = s.find("["), s.rfind("]")
        if i == -1 or j == -1 or j < i:
            return []
        s = s[i:j + 1]
    try:
        data = json.loads(s)
    except Exception:
        return []
    return [d for d in data if isinstance(d, dict)] if isinstance(data, list) else []

def qa_ai_from_chat(api_key, chat_text, model=GEMINI_DEFAULT, images=None, hint=""):
    """把一段蝦皮客服對話（貼上的文字，或多張對話截圖）整理成問題庫格式的「多筆」範本。
    重點：建議回覆以「你實際回過的內容」為準，只做潤飾，不憑空編造承諾（退款、免運等）。
    回傳 (list[dict]，欄位同 QA_COLS 的前五項, usage)。"""
    txt = (chat_text or "").strip()
    imgs = [im for im in (images or []) if im is not None]
    if not txt and not imgs:
        return [], None
    prompt = f"""你是蝦皮賣場 BearJoy 的客服主管 Sharon 的助理。下面是我和客人的客服對話紀錄（可能是文字，也可能是對話截圖）。
請把對話拆解成「一個問題一筆」的客服範本，之後我要靠關鍵字搜尋叫出來直接複製貼給客人。

【規則】
1. 一則對話若含多個不同問題，就拆成多筆；同一個問題重複出現只留一筆。
2. 「建議回覆範本」以我實際回覆過的內容為準，只做語句潤飾與去識別化，**不可以自己編造沒發生過的承諾**（例如退款金額、免運、贈品）。
3. 若對話中我還沒有回覆，才由你依 BearJoy 溫暖專業的語氣補一則合適回覆。
4. 移除客人姓名、電話、地址、訂單編號等個資，改成「您」或〔訂單編號〕之類的佔位字。
5. 「分類」用簡短通用詞，例如：物流、退換貨、商品規格、付款、折價券、售後保固、其他。
6. 「關鍵字」給 3～6 個客人可能會打的字，用空格分隔。
{('7. 額外提示：' + hint.strip()) if hint.strip() else ''}

【輸出格式】只輸出一個 JSON 陣列，不要任何說明文字或程式碼圍欄。每個元素長這樣：
{{"分類":"物流","問題標題":"包裹遲遲未更新","客戶問題範例":"下單三天了物流都沒動…","建議回覆範本":"（可直接貼給客人的完整回覆）","關鍵字":"物流 未更新 出貨"}}

【對話紀錄】
{txt if txt else '（見附上的對話截圖）'}"""
    contents = [prompt] + imgs
    text, usage = gemini_call_costed(api_key, contents, model)
    items = []
    for d in _json_list_from_text(text):
        items.append({
            "分類": str(d.get("分類", "") or "").strip(),
            "問題標題": str(d.get("問題標題", "") or "").strip(),
            "客戶問題範例": str(d.get("客戶問題範例", "") or "").strip(),
            "建議回覆範本": str(d.get("建議回覆範本", "") or "").strip(),
            "關鍵字": str(d.get("關鍵字", "") or "").strip(),
        })
    return [it for it in items if it["問題標題"] or it["客戶問題範例"]], usage

# ==========================================
# 💌 私訊查詢：公開回覆先發、私訊過幾天再回時，用搜尋叫出「當初那則評價 ＋ 要私訊的內容」
#    純查詢，不做任何狀態管理，也完全不寫入雲端（只讀「回覆紀錄」與「評價截圖素材」）。
# ==========================================
DM_SHEET = "回覆紀錄"
DM_BASE_COLS = ["紀錄時間", "客戶帳號", "原始評價內容", "賣場評價回覆", "VIP私訊回覆"]
MAT_SHEET = "評價截圖素材"

def dm_load(doc):
    """唯讀取「回覆紀錄」。回傳 (ws, rows)；rows 由新到舊，每筆含 _row 實際列號。"""
    ws = get_or_create_ws(doc, DM_SHEET)
    values = ws.get_all_values()
    if not values or not any((c or "").strip() for c in values[0]):
        return ws, []
    header = list(values[0])
    rows = []
    for i, r in enumerate(values[1:], start=2):
        if not any((c or "").strip() for c in r):
            continue
        d = {c: (r[j] if j < len(r) else "") for j, c in enumerate(header)}
        for c in DM_BASE_COLS:                 # 舊表缺欄也不會 KeyError
            d.setdefault(c, "")
        d["_row"] = i
        rows.append(d)
    rows.reverse()          # 最新的排最上面
    return ws, rows

def review_shot_find(doc, acc, ts):
    """在「評價截圖素材」找這位客戶當初的評價截圖。
    素材列的第一欄是 f"{年月日_時分秒}_{帳號}|||{規格}"，與回覆紀錄同一個時間戳，先比同秒、
    再退回同帳號最近一張。只讀第一欄（很輕，不會把整批 base64 圖抓下來）。
    回傳 (ws, 列號) 或 (None, None)。"""
    try:
        ws = get_or_create_ws(doc, MAT_SHEET)
        keys = ws.col_values(1)
    except Exception:
        return None, None
    try:
        stamp = datetime.strptime((ts or "").strip(), "%Y-%m-%d %H:%M:%S").strftime("%Y%m%d_%H%M%S")
    except Exception:
        stamp = ""
    tail = "_" + (acc or "").strip()
    same_acc = []
    for i, k in enumerate(keys, start=1):
        head = str(k or "").split("|||")[0].strip()
        if not head:
            continue
        if not head.endswith(tail):
            continue
        if stamp and head.startswith(stamp):
            return ws, i                      # 同秒＝就是這則評價的原圖
        same_acc.append(i)
    return (ws, same_acc[-1]) if same_acc else (None, None)

def review_shot_load(ws, row):
    """把素材那一列的 base64 切塊還原成圖片（只抓那一列）。失敗回 None。"""
    try:
        vals = ws.row_values(row)
        return base64_chunks_to_img([c for c in vals[1:] if c])
    except Exception:
        return None

# 💰 全程式共用的花費累計（任何用到 API 的功能都呼叫這兩個，介面就有一致的費用顯示）
def ai_track_cost(usage):
    """把一次 API 用量累計進工作階段總花費，並記為最近一次。usage 為 gemini_call_costed 回傳的 dict。"""
    if not usage:
        return
    st.session_state["ai_last_usage"] = usage
    st.session_state["ai_cost_total"] = st.session_state.get("ai_cost_total", 0.0) + usage.get("cost", 0.0)

def ai_render_cost(model_key=None):
    """在目前位置畫出花費資訊：上次用量＋本次開啟累計（美金估算）。model_key 可順帶顯示目前模型。"""
    last = st.session_state.get("ai_last_usage")
    total = st.session_state.get("ai_cost_total", 0.0)
    if last:
        free = "（免費額度內，實際 $0）" if last.get("free") else "（付費模型，會實際扣費）"
        used_lbl = GEMINI_PRICES.get(last.get("model", ""), {}).get("label", last.get("model", ""))
        fb = "　⚠️ 你選的模型打不通，已自動改用此款" if last.get("fellback") else ""
        st.caption(f"💰 上次生成（{used_lbl}）：輸入 {last['in']} ＋ 輸出 {last['out']} tokens"
                   f"｜估算 US${last['cost']:.5f} {free}{fb}")
    else:
        st.caption("💰 上次生成：尚未使用 AI")
    tail = f"　({GEMINI_PRICES[model_key]['label']})" if model_key in GEMINI_PRICES else ""
    st.caption(f"📊 本次開啟累計估算：US${total:.5f}{tail}")

# 顧客評價原圖要備份到的 Google Drive 資料夾
DRIVE_FOLDER_ID = "1ZamXtEG9tiG6HTQJXTD6e_am6u3B4bGz"

# ☁️ Drive 備份改走「Apps Script 中轉」：個人 Gmail 的服務帳號沒有儲存空間（會 403），
#    改由使用者自己部署的 Apps Script 網頁應用程式，以「使用者本人身分」把圖存進 Drive。
#    🔐 網址＋密碼不寫死在程式碼（避免公開 repo 外洩）：雲端讀 st.secrets、本機讀 drive_config.txt。
def _load_drive_relay_cfg():
    try:
        u = st.secrets.get("apps_script_drive_url", "")
        s = st.secrets.get("apps_script_drive_secret", "")
        if u and s:
            return u, s
    except Exception:
        pass
    try:
        p = os.path.join(os.path.dirname(__file__), "drive_config.txt")
        if os.path.exists(p):
            ls = open(p, encoding="utf-8").read().splitlines()
            if len(ls) >= 2:
                return ls[0].strip(), ls[1].strip()
    except Exception:
        pass
    return "", ""

APPS_SCRIPT_DRIVE_URL, APPS_SCRIPT_DRIVE_SECRET = _load_drive_relay_cfg()

@st.cache_resource(show_spinner=False)
def _drive_service():
    """建立 Google Drive API 連線（用與試算表相同的服務帳號金鑰），快取重用。"""
    scope = ["https://www.googleapis.com/auth/drive"]
    creds = None
    try:
        if "type" in st.secrets:
            creds = ServiceAccountCredentials.from_json_keyfile_dict(dict(st.secrets), scope)
    except Exception:
        pass
    if not creds:
        kp = os.path.join(os.path.dirname(__file__), "google_key.json")
        if os.path.exists(kp):
            creds = ServiceAccountCredentials.from_json_keyfile_name(kp, scope)
    if not creds:
        raise RuntimeError("找不到金鑰")
    import httplib2
    from googleapiclient.discovery import build
    return build("drive", "v3", http=creds.authorize(httplib2.Http()), cache_discovery=False)

def _safe_filename(s):
    s = re.sub(r'[\\/:*?"<>|\n\r\t]+', " ", str(s)).strip()
    return (s[:80] or "未命名")

def section_block(emoji, title, desc=""):
    """大區塊標題：左色條標題＋旁邊原生「?」說明（hover 顯示，與『保存截圖』的 ? 一致）。"""
    st.subheader(f"{emoji} {title}", help=(desc or None), anchor=False)

def upload_img_to_drive(img, filename, folder_id=DRIVE_FOLDER_ID):
    """把 PIL 圖片透過 Apps Script 中轉，以使用者本人身分存進 Drive 資料夾；
    回傳 (True, 連結) 或 (False, 錯誤訊息)。"""
    import requests
    if not APPS_SCRIPT_DRIVE_URL:
        return False, "尚未設定 Apps Script 中轉網址（雲端要在 secrets 設 apps_script_drive_url／本機要有 drive_config.txt）"
    # 📉 先縮到 1600px、轉 JPEG：實測 1.2MB 的 PNG 會變 200KB 左右，上傳快很多也不容易逾時
    im = img.convert("RGB")
    im.thumbnail((1600, 1600), Image.Resampling.LANCZOS)
    buf = BytesIO()
    im.save(buf, format="JPEG", quality=92, optimize=True)
    b64 = base64.b64encode(buf.getvalue()).decode()
    fn = re.sub(r"\.(png|jpeg|jpg)$", "", str(filename), flags=re.I) + ".jpg"
    last = ""
    for attempt in range(2):        # 失敗自動重試一次（網路偶發、Apps Script 忙碌）
        try:
            r = requests.post(APPS_SCRIPT_DRIVE_URL, json={
                "secret": APPS_SCRIPT_DRIVE_SECRET,
                "filename": fn,
                "mimeType": "image/jpeg",
                "dataBase64": b64,
            }, timeout=90)
            data = r.json()
            if data.get("ok"):
                return True, data.get("url")
            last = str(data.get("err", "上傳失敗"))
        except Exception as e:
            last = f"{type(e).__name__}: {e}"
        time.sleep(2)
    return False, last or "上傳失敗"

def _review_spec(content):
    """從評價內容抓出『規格』。整串都算規格（含中括號與顏色），不是只有【】內的字。
    容錯：半形或全形冒號（規格: / 規格：）、規格可能不在第一行（跨行搜尋）。
    取『規格:』後面到第一個分隔符（｜ / | / 逗號 / 換行）為止的完整文字。抓不到回空字串。"""
    s = str(content)
    m = re.search(r'規格\s*[：:]\s*([^\n]+)', s)
    if not m:
        return ""
    val = m.group(1)
    # 切到最早出現的分隔符（避免把後面重複或別欄的字也吃進來）
    idxs = [val.find(c) for c in ["｜", "|", "，", ","] if c in val]
    if idxs:
        val = val[:min(idxs)]
    return _spec_no_color(val)

def _spec_no_color(val):
    """規格去顏色：款式多在中括號 [..] 內，括號後通常是顏色 → 截到最後一個 ] 為止，
    讓「[三層款…]黑色」「[三層款…]粉色」合併成同一款，不會分太細。沒有中括號就原樣。"""
    val = str(val).strip().rstrip("，,、 ")
    if "]" in val:
        val = val[:val.rfind("]") + 1]
    return val[:40]

def _mat_meta(row_id):
    """解析評價截圖素材 id：『YYYYMMDD_HHMMSS_帳號|||規格』→ (日期, 帳號, 規格)。舊資料沒有規格回空。"""
    s = str(row_id)
    spec = ""
    if "|||" in s:
        s, spec = s.split("|||", 1)
    parts = s.split("_")
    date = parts[0] if parts else ""
    acc = "_".join(parts[2:]) if len(parts) > 2 else (parts[-1] if parts else "")
    return date, acc, spec.strip()

def _excel_align_left_top(ws):
    """openpyxl 對齊：依表頭判斷，日期/時間/天數/次數欄靠右，其餘靠左；皆靠上＋自動換行。"""
    try:
        from openpyxl.styles import Alignment
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        right_cols = {i + 1 for i, h in enumerate(headers) if _col_align_right(h)}
        for row in ws.iter_rows():
            for c in row:
                horiz = "right" if c.column in right_cols else "left"
                c.alignment = Alignment(horizontal=horiz, vertical="top", wrap_text=True)
    except Exception:
        pass

# ==========================================
# ✨ 銷售加值工具：AI 分析、好評圖
# ==========================================
def gemini_generate(api_key, contents, models=None):
    """呼叫 Gemini，自動換模型重試；失敗回傳 None。
    ✨ 速度優化：第一次嘗試不再空等 2 秒，只有「失敗要重試」時才退避等待
    （等待時間隨次數加長，避免觸發免費版流量限制）。
    contents 可為純文字 [prompt]，也可為圖文 [prompt, img]，批次處理共用同一套邏輯。"""
    try:
        client = genai.Client(api_key=api_key)
    except Exception:
        return None
    models = models or ["gemini-2.5-flash", "gemini-2.0-flash", "gemini-1.5-flash"]
    first = True
    for model_name in models:
        for attempt in range(3):
            try:
                if not first:
                    time.sleep(2 + attempt * 2)  # 退避：第 0/1/2 次重試分別等 2/4/6 秒
                first = False
                resp = client.models.generate_content(model=model_name, contents=contents)
                return resp.text
            except Exception:
                continue
    return None


def _extract_section(text, start_tag, end_tags):
    """從 AI 回覆中安全擷取某段落，找不到標籤回傳 None（取代易出錯的多重 split）。"""
    if not text or start_tag not in text:
        return None
    seg = text.split(start_tag, 1)[1]
    for et in end_tags:
        if et in seg:
            seg = seg.split(et, 1)[0]
            break
    return seg.strip()

def _mask_account(acc):
    """遮罩客戶帳號中間字元，保護隱私（公開貼圖用）。"""
    acc = str(acc).strip()
    if len(acc) <= 2:
        return acc[:1] + "*"
    return acc[0] + "*" * (len(acc) - 2) + acc[-1]

def _clean_text(s):
    """只保留中文、英數、常用標點，移除 emoji / 符號 / 亂碼方框。"""
    out = []
    for ch in str(s):
        o = ord(ch)
        if ch in "\n\t ":
            out.append(ch)
        elif 0x20 <= o <= 0x7E:
            out.append(ch)
        elif 0x2018 <= o <= 0x201F:
            out.append(ch)
        elif 0x3000 <= o <= 0x303F:
            out.append(ch)
        elif 0x3400 <= o <= 0x4DBF:
            out.append(ch)
        elif 0x4E00 <= o <= 0x9FFF:
            out.append(ch)
        elif 0xFF00 <= o <= 0xFFEF:
            out.append(ch)
    return "".join(out).strip()

def _wrap_cjk(text, max_chars):
    text = " ".join(_clean_text(text).split())
    lines, cur = [], ""
    for ch in text:
        cur += ch
        if len(cur) >= max_chars:
            lines.append(cur); cur = ""
    if cur:
        lines.append(cur)
    return lines or [""]

def _text_w(draw, text, font):
    try:
        b = draw.textbbox((0, 0), text, font=font)
        return b[2] - b[0]
    except Exception:
        return len(text) * 20

def _render_content(reviews, template):
    """回傳一張內容圖（米底），之後再縮放置中到目標尺寸。
    ✨ 大標題置中；每則評價的星星、帳號、內文一律『靠左對齊』；規格已併入評價內容一起顯示。
    ✨ 卡片高度用『實際量測的文字高度』決定，文字一定包在框內、不會超出。"""
    BW, pad = 1000, 50
    stars = "★ ★ ★ ★ ★"
    is_quote = (template == "quote")
    dummy = ImageDraw.Draw(Image.new("RGB", (1, 1)))

    title_font = get_chinese_font(52 if is_quote else 50)
    body_font = get_chinese_font(40 if is_quote else 34)
    acc_font = get_chinese_font(28)
    star_font = get_chinese_font(30)

    line_gap = 12          # 內文行距
    pads_v = 28            # 卡片上下內留白
    star_h, acc_h = 44, 42
    gap_between = 30       # 卡片之間距
    max_chars = 18 if is_quote else 22

    blocks = [(_mask_account(it[0]), "\n".join(_wrap_cjk(it[1], max_chars))) for it in reviews]

    def measure(text, font):
        try:
            bb = dummy.multiline_textbbox((0, 0), text, font=font, spacing=line_gap)
            return bb[2] - bb[0], bb[3] - bb[1]
        except Exception:
            return 600, (text.count("\n") + 1) * 50

    heights = []
    for _, body in blocks:
        _, bh = measure(body, body_font)
        heights.append(pads_v + star_h + acc_h + 10 + bh + pads_v)

    title = "顧客真實心得" if is_quote else "BearJoy 顧客真實好評"
    top = 150
    H = top + sum(h + gap_between for h in heights) + pad
    img = Image.new("RGB", (BW, H), "#FAF8F5")
    d = ImageDraw.Draw(img)
    d.text(((BW - _text_w(d, title, title_font)) / 2, 55), title, font=title_font, fill="#4A4238")

    y = top
    for (acc, body), h in zip(blocks, heights):
        cb = y + h
        if not is_quote:
            try:
                d.rounded_rectangle([pad, y, BW - pad, cb], radius=22, fill="#FFFFFF", outline="#E6E2D8", width=2)
            except Exception:
                d.rectangle([pad, y, BW - pad, cb], fill="#FFFFFF", outline="#E6E2D8")
        x0 = pad + 36           # 卡片左內留白
        ty = y + pads_v
        # 靠左：星星 + 帳號 + 內文（大標題仍置中，於上方已畫）
        d.text((x0, ty), stars, font=star_font, fill="#E0A96D")
        d.text((x0, ty + star_h), f"@{acc}", font=acc_font, fill="#A0998C")
        d.multiline_text((x0, ty + star_h + acc_h + 10), body,
                         font=body_font, fill="#4A4238", spacing=line_gap, align="left")
        y = cb + gap_between
    return img

def make_review_image(reviews, size=(1080, 1080), template="cards"):
    """reviews: [(帳號, 內容), ...]；回傳指定尺寸 (W,H) 的 PIL Image。"""
    TW, TH = size
    content = _render_content(reviews, template)
    canvas = Image.new("RGB", (TW, TH), "#FAF8F5")
    margin = int(min(TW, TH) * 0.04)
    avail_w, avail_h = TW - margin * 2, TH - margin * 2
    scale = avail_w / content.width
    if content.height * scale > avail_h:
        scale = avail_h / content.height
    nw, nh = max(1, int(content.width * scale)), max(1, int(content.height * scale))
    content_r = content.resize((nw, nh), Image.LANCZOS)
    canvas.paste(content_r, ((TW - nw) // 2, (TH - nh) // 2))
    return canvas

def _render_collage(images):
    """把實際上傳的評價截圖拼成一張內容圖（米底＋標題＋masonry 拼貼）。
    ✨ 滿版＋高解析：縮小邊距讓評價圖盡量填滿、BW 拉高讓文字/截圖更清晰。"""
    # ✨ 畫質升級：內部渲染整體放大 1.5×（版面比例不變、只是更密），縮到目標尺寸後截圖更銳利、不糊
    BW, pad, gap = 2400, 27, 21
    title_font = get_chinese_font(100)
    star_font = get_chinese_font(70)
    top = 345  # 標題＋五星留足空間，圖片從這裡才開始（標題與星星不擁擠）
    cols = 2 if len(images) > 1 else 1
    cell_w = int((BW - 2 * pad - (cols - 1) * gap) / cols)
    placed, col_y = [], [top] * cols
    for im in images:
        im = im.convert("RGB")
        sw = cell_w
        sh = max(1, int(im.height * sw / im.width))
        c = col_y.index(min(col_y))
        x = pad + c * (cell_w + gap)
        y = col_y[c]
        placed.append((im.resize((sw, sh), Image.LANCZOS), x, y, sw, sh))
        col_y[c] += sh + gap
    H = max(col_y) + pad - gap
    img = Image.new("RGB", (BW, H), "#FAF8F5")
    d = ImageDraw.Draw(img)
    title, stars = "BearJoy 顧客真實好評", "★ ★ ★ ★ ★"
    d.text(((BW - _text_w(d, title, title_font)) / 2, 63), title, font=title_font, fill="#4A4238")
    d.text(((BW - _text_w(d, stars, star_font)) / 2, 225), stars, font=star_font, fill="#E0A96D")
    for im, x, y, sw, sh in placed:
        try:
            d.rounded_rectangle([x - 5, y - 5, x + sw + 5, y + sh + 5], radius=18, outline="#E6E2D8", width=5)
        except Exception:
            d.rectangle([x - 5, y - 5, x + sw + 5, y + sh + 5], outline="#E6E2D8")
        img.paste(im, (x, y))
    return img

def make_collage_image(images, size=(1080, 1080)):
    """把實際截圖拼貼後縮放置中到目標尺寸。"""
    TW, TH = size
    content = _render_collage(images)
    canvas = Image.new("RGB", (TW, TH), "#FAF8F5")
    margin = int(min(TW, TH) * 0.015)  # 邊距縮小，評價圖更滿版
    aw, ah = TW - margin * 2, TH - margin * 2
    scale = aw / content.width
    if content.height * scale > ah:
        scale = ah / content.height
    nw, nh = max(1, int(content.width * scale)), max(1, int(content.height * scale))
    canvas.paste(content.resize((nw, nh), Image.LANCZOS), ((TW - nw) // 2, (TH - nh) // 2))
    return canvas

def _draw_marker(img, x, y):
    """在預覽圖畫紅色十字準心，標出文字中心點。只用於畫面預覽，不會存進實際檔案。
    拉動「左右 / 上下」拉桿時，這個十字會即時跟著移動，方便對準想要的位置。"""
    im = img.convert("RGB").copy()
    d = ImageDraw.Draw(im)
    r = max(16, int(min(im.width, im.height) * 0.04))
    # 先畫白色粗框（深色底圖也看得見），再疊紅色細線
    d.line([(x - r, y), (x + r, y)], fill="#FFFFFF", width=6)
    d.line([(x, y - r), (x, y + r)], fill="#FFFFFF", width=6)
    d.line([(x - r, y), (x + r, y)], fill="#E0533D", width=2)
    d.line([(x, y - r), (x, y + r)], fill="#E0533D", width=2)
    d.ellipse([x - 5, y - 5, x + 5, y + 5], outline="#E0533D", width=2)
    return im

def _stamp_coupon(base_img, text, color, size, cx, cy, rot):
    """在底圖 (cx,cy) 壓上旋轉文字，回傳 (RGB圖, 文字寬, 文字高)。壓印預覽與儲存共用。"""
    preview = base_img.copy().convert("RGBA")
    font = get_chinese_font(size)
    dd = ImageDraw.Draw(Image.new('RGBA', (1, 1)))
    ox = oy = 0
    try:
        bb = dd.multiline_textbbox((0, 0), text, font=font, align="center")
        ox, oy = bb[0], bb[1]            # 墨水相對「畫字錨點(上緣)」的左/上偏移
        tw, th = bb[2] - bb[0], bb[3] - bb[1]
    except Exception:
        tw, th = 200, 100
    lw, lh = max(2, int(tw * 2.5)), max(2, int(th * 2.5))
    layer = Image.new('RGBA', (lw, lh), (255, 255, 255, 0))
    ld = ImageDraw.Draw(layer)
    # 🔑 對位修正：multiline_text 的錨點是「上緣/ascender」而非實際墨水頂端，
    # 只用 th/2 置中會讓字往下偏 oy（中文/數字尤其明顯）。扣掉 (ox,oy) 後，
    # 「墨水中心」才會真正落在圖層中心 → 壓出來對準畫布上拖到的位置，不再偏下。
    px = lw / 2 - tw / 2 - ox
    py = lh / 2 - th / 2 - oy
    try:
        ld.multiline_text((px, py), text, fill=color, font=font, align="center")
    except Exception:
        ld.text((px, py), text, fill=color, font=font)
    rl = layer.rotate(-rot, expand=True, resample=Image.BICUBIC)
    preview.alpha_composite(rl, (int(cx - rl.width / 2), int(cy - rl.height / 2)))
    return preview.convert("RGB"), tw, th

def _parse_coupset(row):
    """解析壓印記憶字串 x|y|size|rot|color|text（舊資料沒有最後的 text 也吃得下）。"""
    if not (row and len(row) > 1):
        return {}
    p = str(row[1]).split("|")
    try:
        return {"x": int(float(p[0])), "y": int(float(p[1])),
                "size": int(float(p[2])), "rot": int(float(p[3])),
                "color": p[4] if len(p) > 4 else "#FFFFFF",
                "text": p[5] if len(p) > 5 else ""}
    except Exception:
        return {}

def threaded_update_order(creds_dict, sheet_url, order_str):
    try:
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        if creds_dict:
            creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
        else:
            key_path = os.path.join(os.path.dirname(__file__), "google_key.json")
            creds = ServiceAccountCredentials.from_json_keyfile_name(key_path, scope)
        gc = gspread.authorize(creds)
        doc = gc.open_by_url(sheet_url)
        ws = doc.worksheet("系統設定")
        all_vals = ws.col_values(1)
        if "coupon_order" in all_vals:
            idx = all_vals.index("coupon_order") + 1
            ws.update_cell(idx, 2, order_str)
        else:
            ws.append_row(["coupon_order", order_str])
    except Exception: pass

def trigger_order_save(url, active_slots):
    creds_dict = None
    try:
        if "type" in st.secrets:
            creds_dict = dict(st.secrets)
    except Exception: pass
    order_str = ",".join(map(str, active_slots))
    threading.Thread(target=threaded_update_order, args=(creds_dict, url, order_str)).start()

BATCH_SYS_PROMPT = """
                    你是一個專業的蝦皮賣場客服主管 Sharon。請精準辨識圖片中的 [ACCOUNT] 帳號與 [REVIEW] 內容。
                    
                    【極度嚴格規定】
                    1. 字數與行數「必須」與下方範例長度一致，絕不可自行長篇大論！
                    2. 務必使用分段與換行符號，保持版面清爽舒適。
                    3. 完全模仿 BearJoy Sharon 的溫暖語氣，並加上適當的 Emoji。
                    4. 【絕對要客製化】：賣場回覆與私訊回覆中，必須「100% 精準引用」客人實際寫出的優點關鍵字。
                    
                    [範例評價]：非常好用，厚的材質...回購第二次
                    [範例賣場回覆]：
                    親愛的顧客您好，

                    感謝您的五星好評與再次回購！🌟
                    很高興厚實的材質能讓您覺得方便。👶
                    謝謝您肯定我們的品質，
                    非常珍貴像您這樣用心回饋的好顧客！❤️
                    期待未來能繼續為您服務。

                    —— BearJoy Sharon

                    [範例私訊回覆]：
                    親愛的 (客戶帳號)，

                    感謝您撥空寫下如此詳細的評價！
                    看到您滿意我們的厚實材質與實用性，我們非常感動。🥰
                    為了表達萬分謝意，已將您列入VIP客戶名單！✨
                    未來若有專屬優惠定會第一時間發送給您。🎁
                    再次感謝支持！

                    —— BearJoy Sharon

                    請依照以下標籤輸出：
                    [ACCOUNT]
                    (客戶帳號)
                    [REVIEW]
                    (評價內容)
                    [SPEC]
                    (顧客購買的規格/款式：蝦皮評價幾乎都會在帳號下方或評價區顯示「規格」「分類」或商品變體，務必逐字「完整讀出」，例如「[三層款 可拆卸 十合一]黑色」「[旗艦款 壓縮 七合一]」「筷子瀝水架三格掛勾款[304接水盤]」等，含中括號與顏色都要。先在整張圖找有沒有「規格」「分類」字樣，找到就照抄。只有當圖片真的完全找不到任何規格/分類/款式字樣時，才寫「無」。不可漏抓。)
                    [RTIME]
                    (這則評價本身的日期時間：蝦皮評價的帳號旁或評價下方會顯示，例如「2026-08-20 14:32」或「2026-08-20」。照原樣抄出數字即可，只有真的找不到才寫「無」。)
                    [PUBLIC]
                    (賣場評價回覆)
                    [PRIVATE]
                    (私訊回覆)
"""

# ==========================================
# 🏃 批次評價「背景執行」引擎（手機滑掉畫面／螢幕暗掉也不會中斷）
#    原因：Streamlit 每次執行都綁在瀏覽器連線上，手機把分頁切走或螢幕關掉、連線一斷，
#          正在跑的那次執行就會被中止 → AI 寫到一半就沒了。
#    解法：把整批工作丟進背景執行緒（裡面完全不碰 st.*，所以連線斷了照跑），
#          進度與結果寫進 batch_job_store()（@st.cache_resource＝整個程式共用，換頁、重新整理都還在），
#          畫面只負責每 2 秒讀一次進度，顯示「已完成 X／共 Y 筆，還剩 Z 筆」。
# ==========================================

def _rev_norm(t):
    """評價內容正規化：拿掉「規格：」字樣與所有空白標點，只留文字本體。
    AI 每次讀同一張圖，標點與規格前綴常常不一樣，正規化後才比得出是同一則。"""
    t = re.sub(r"規格\s*[:：]", "", str(t))
    return re.sub(r"[\s\W_]+", "", t, flags=re.UNICODE)


def _rtime_norm(t):
    """評價時間正規化：只留數字（2026-08-20 14:32 → 202608201432），格式寫法不同也比得出來。"""
    return re.sub(r"\D", "", str(t or ""))


def _same_review(a_rev, a_rt, b_rev, b_rt, thr=0.8):
    """是不是同一則評價：兩邊都讀得到評價時間就直接比時間（最準）；
    否則退回比內容相似度（門檻 0.8）。"""
    ra, rb = _rtime_norm(a_rt), _rtime_norm(b_rt)
    if len(ra) >= 8 and len(rb) >= 8:
        return ra == rb
    if not a_rev or not b_rev:
        return False
    return difflib.SequenceMatcher(None, a_rev, b_rev).ratio() >= thr


@st.cache_resource(show_spinner=False)
def batch_job_store():
    """跨工作階段共用的批次任務儲存區：程式沒關就在，手機切走再回來也讀得到同一份進度。"""
    return {"lock": threading.Lock(), "job": None}


def batch_job_get():
    return batch_job_store().get("job")


def _batch_open_doc(creds_dict, sheet_url):
    """背景執行緒專用：自己開一條 Google 試算表連線（不可共用畫面那條快取連線）。"""
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    if creds_dict:
        creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
    else:
        creds = ServiceAccountCredentials.from_json_keyfile_name(
            os.path.join(os.path.dirname(__file__), "google_key.json"), scope)
    return gspread.authorize(creds).open_by_url(sheet_url)


def _batch_sync_cloud(job, doc_bg, results):
    """把整批結果寫進「回覆紀錄」並更新 VIP 名單；同時把雲端列號回填給畫面（供逐筆改寫用）。"""
    try:
        ws_history = get_or_create_ws(doc_bg, "回覆紀錄")
        existing = ws_history.get_all_values()
        header = ["紀錄時間", "客戶帳號", "原始評價內容", "賣場評價回覆", "VIP私訊回覆", "評價時間"]
        if len(existing) == 0:
            ws_history.append_row(header)
            existing = [header]
        elif len(existing[0]) < 6:
            # 舊表只有 5 欄 → 補上第 6 欄「評價時間」（舊資料留白，不影響既有內容）
            try:
                ws_history.update_cell(1, 6, "評價時間")
            except Exception:
                pass

        # 🔁 去重（升級版）：同一帳號 ＋ 同一個「評價時間」＝同一則評價；讀不到時間才比內容相似度。
        #    命中就【覆蓋更新原本那一列】（後跑的是最終版），不再新增重複列，
        #    VIP 互動次數也不會被重複加。
        _idx = {}      # 帳號 → [(列號或 None, 正規化內容, 評價時間)]
        for _r0, _rv in enumerate(existing[1:], start=2):
            if len(_rv) > 2:
                _idx.setdefault(str(_rv[1]).strip().lower(), []).append(
                    (_r0, _rev_norm(_rv[2]), (_rv[5] if len(_rv) > 5 else "")))
        new_rows, dup_count, upd_count, row_map = [], 0, 0, {}
        for row in results:
            _acc_k = str(row[1]).strip().lower()
            _nv, _nt = _rev_norm(row[2]), (row[5] if len(row) > 5 else "")
            _hit, _self_dup = None, False
            for _rn, _ov, _ot in _idx.get(_acc_k, []):
                if _same_review(_ov, _ot, _nv, _nt):
                    if _rn is None:
                        _self_dup = True     # 同一批裡自己重複（同一則評價上傳了兩張截圖）
                    else:
                        _hit = _rn
                    break
            if _self_dup:
                dup_count += 1
                continue
            if _hit:
                try:
                    ws_history.update(f"A{_hit}:F{_hit}", [row])
                    upd_count += 1
                except Exception:
                    dup_count += 1
                row_map[(str(row[1]).strip(), str(row[2]).strip())] = _hit
                continue
            _idx.setdefault(_acc_k, []).append((None, _nv, _nt))
            new_rows.append(row)
        if new_rows:
            _first_new = len(existing) + 1      # append 前的總列數＋1＝第一筆新資料的列號
            ws_history.append_rows(new_rows)
            for _i, _nr in enumerate(new_rows):
                row_map[(str(_nr[1]).strip(), str(_nr[2]).strip())] = _first_new + _i
        for it in list(job.get("items", [])):
            it["row"] = row_map.get((str(it["acc"]).strip(), str(it["rev"]).strip()))

        ws_vip = get_or_create_ws(doc_bg, "VIP名單")
        vip_vals = ws_vip.get_all_values()
        if len(vip_vals) == 0:
            v_header = ["客戶帳號", "首次互動", "最後互動", "互動次數"]
            ws_vip.append_row(v_header)
            vip_vals = [v_header]
        v_header = vip_vals[0]
        vip_records = [dict(zip(v_header, r)) for r in vip_vals[1:]]
        date_str = datetime.now().strftime("%Y-%m-%d")
        for row in new_rows:
            account = row[1]
            if account == "未知":
                continue
            fi = next((i for i, r in enumerate(vip_records)
                       if str(r.get('客戶帳號', '')) == account), -1)
            if fi != -1:
                ws_vip.update_cell(fi + 2, 3, date_str)
                ws_vip.update_cell(fi + 2, 4, int(vip_records[fi].get('互動次數', 0)) + 1)
            else:
                ws_vip.append_row([account, date_str, date_str, 1])
        try:
            ws_vip.format("A:A", {"horizontalAlignment": "LEFT", "verticalAlignment": "TOP"})
            ws_vip.format("B:D", {"horizontalAlignment": "RIGHT", "verticalAlignment": "TOP"})
        except Exception:
            pass
        msg = f"🎉 完美同步！新增 {len(new_rows)} 筆紀錄"
        if upd_count:
            msg += f"｜{upd_count} 筆是同一則評價重跑，已直接更新成最新版本（不新增重複列）"
        if dup_count:
            msg += f"｜略過 {dup_count} 筆重複評價，不重複計算互動次數"
        job["msg"] = msg
    except Exception as e:
        job["notes"].append(f"⚠️ 雲端同步失敗（結果還在畫面上）：{str(e)[:120]}")


def _batch_worker(job, imgs, opts):
    """背景執行緒本體：一張一張叫 AI、存素材、備份 Drive，最後整批寫回雲端。全程不碰 st.*。"""
    lock = batch_job_store()["lock"]
    try:
        prompt_base = BATCH_SYS_PROMPT + ("\n注意：此為二回購老客，請加入朋友般的尊榮感。"
                                          if opts.get("is_vip") else "")
        if opts.get("batch_ins"):
            prompt_base += ("\n【本批額外要求（務必遵守）】：" + opts["batch_ins"] +
                            "\n請把這個要求自然融入 [PUBLIC] 與 [PRIVATE] 兩段回覆中，"
                            "不要另外條列說明，也不可以自行新增原本沒有的優惠或承諾。")
        doc_bg = None
        try:
            doc_bg = _batch_open_doc(opts.get("creds_dict"), opts.get("sheet_url"))
        except Exception as e:
            job["notes"].append(f"⚠️ 雲端連線失敗，這批只會顯示在畫面、不寫回雲端：{str(e)[:90]}")

        results = []
        for i, (fname, img) in enumerate(imgs):
            job["current_name"] = fname
            # ✨ 第一張立即處理，之後每張間隔 4 秒，避免免費版流量限制
            if i > 0:
                time.sleep(4)
            res_text, usage = gemini_call_costed(opts["api_key"], [prompt_base, img], opts["model"])
            if usage:
                job["cost"] = job.get("cost", 0.0) + usage.get("cost", 0.0)
                job["last_usage"] = usage
            if not res_text:
                job["notes"].append(f"⚠️ 檔案 {fname} 處理失敗（AI 沒回應），已略過。")
                job["done"] = i + 1
                continue

            acc = _extract_section(res_text, "[ACCOUNT]", ["[REVIEW]"]) or "未知"
            rev = _extract_section(res_text, "[REVIEW]", ["[SPEC]", "[PUBLIC]"]) or "解析失敗"
            spec = (_extract_section(res_text, "[SPEC]", ["[RTIME]", "[PUBLIC]"]) or "").strip()
            # 🕒 評價本身的日期時間：拿它當「這則評價的身分證」來去重，比比對文字準
            rtime = (_extract_section(res_text, "[RTIME]", ["[PUBLIC]"]) or "").strip()
            if rtime in ("無", "None", "-"):
                rtime = ""
            pub = _extract_section(res_text, "[PUBLIC]", ["[PRIVATE]"]) or "解析失敗"
            priv = _extract_section(res_text, "[PRIVATE]", []) or "解析失敗"
            if spec and spec != "無" and rev != "解析失敗" and not rev.startswith("規格"):
                rev = f"規格：{spec}｜{rev}"
            if priv != "解析失敗" and opts.get("repurchase_code"):
                offer_txt = f"（{opts['repurchase_offer']}）" if opts.get("repurchase_offer") else ""
                priv = priv + (f"\n\nP.S. 送您專屬回購碼 👉 {opts['repurchase_code']}{offer_txt}"
                               f"\n下次下單輸入即可享優惠，期待再為您服務 🎁")
            now = datetime.now()

            # 💾 保存原始評價截圖到雲端，供日後做素材
            if opts.get("save_screenshots") and doc_bg is not None:
                try:
                    ws_mat = get_or_create_ws(doc_bg, "評價截圖素材")
                    ws_mat.append_row([f"{now.strftime('%Y%m%d_%H%M%S')}_{acc}|||{_review_spec(rev)}"]
                                      + img_to_chunks_compact(img.copy()))
                except Exception as e:
                    job["notes"].append(f"⚠️ {acc} 的截圖素材保存略過（不影響回覆）：{str(e)[:80]}")

            # ☁️ 備份原圖到 Google Drive 資料夾，檔名＝「日期 評價圖-規格」
            if opts.get("save_to_drive"):
                # 檔名＝「日期 評價圖-規格_帳號」：同一款規格常有十幾位客人，
                # 只用規格會全部同名分不出來，所以末尾補上帳號。
                _spec_for_name = _review_spec(rev) or (spec if spec and spec != "無" else acc)
                _fname = (f"{now.strftime('%Y%m%d')} 評價圖-{_safe_filename(_spec_for_name)}"
                          f"_{_safe_filename(acc)}.png")
                ok_d, info_d = upload_img_to_drive(img.copy(), _fname)
                if ok_d:
                    job["drive_ok"] = job.get("drive_ok", 0) + 1
                else:
                    job["drive_fail"] = job.get("drive_fail", 0) + 1
                    # 失敗原因存在任務裡（不是只印在畫面），連線斷掉也不會不見
                    job["notes"].append(f"⚠️ {acc}（{_fname}）未能備份到 Drive：{str(info_d)[:110]}")

            with lock:
                job["items"].append({"acc": acc, "rev": rev, "pub": pub, "priv": priv,
                                     "rtime": rtime, "row": None})
            results.append([now.strftime("%Y-%m-%d %H:%M:%S"), acc, rev, pub, priv, rtime])
            job["done"] = i + 1

        if doc_bg is not None and results:
            _batch_sync_cloud(job, doc_bg, results)
        job["finished"] = datetime.now().strftime("%H:%M:%S")
        job["status"] = "done"
    except Exception as e:
        job["error"] = str(e)
        job["status"] = "error"


def batch_job_start(imgs, opts):
    """建立一個新的批次任務，丟到背景執行緒跑，立刻回傳（畫面不會被卡住）。"""
    store = batch_job_store()
    job = {"id": datetime.now().strftime("%Y%m%d%H%M%S"), "status": "running",
           "drive_ok": 0, "drive_fail": 0,
           "total": len(imgs), "done": 0, "current_name": "", "items": [], "notes": [],
           "msg": "", "error": "", "cost": 0.0, "last_usage": None,
           "started": datetime.now().strftime("%H:%M:%S"), "finished": ""}
    store["job"] = job
    threading.Thread(target=_batch_worker, args=(job, imgs, opts), daemon=True).start()
    return job


@st.fragment(run_every=2)
def render_batch_progress():
    """每 2 秒自動更新的進度區：顯示到第幾筆、還剩幾筆；已跑完的先秀出來可以先複製。"""
    job = batch_job_get()
    if not job:
        return
    total, done = int(job.get("total") or 0), int(job.get("done") or 0)
    if job.get("status") != "running":
        st.rerun(scope="app")       # 跑完 → 整頁重跑，改由下方結果區呈現
        return
    st.progress(min(1.0, done / total) if total else 0.0,
                text=f"🏃 AI 撰寫中… 已完成 {done} / 共 {total} 筆，還剩 {max(0, total - done)} 筆")
    st.caption(f"⏱ {job.get('started','')} 開始｜正在處理：{job.get('current_name','')}")
    st.info("📱 可以關螢幕或切去別的 App，這批會在背景繼續跑；回來重開這頁就會接上進度。")
    for _n in list(job.get("notes", []))[-3:]:
        st.caption(_n)
    for _i, it in enumerate(list(job.get("items", []))):
        with st.expander(f"✅ 第 {_i + 1} 筆　客戶帳號：{it['acc']}", expanded=False):
            st.markdown(f"**📝 原始評價內容:** {it['rev']}")
            st.markdown("**📢 賣場回覆 (點擊右上角複製):**")
            st.code(it["pub"], language="text")
            st.markdown("**💌 私訊回覆 (點擊右上角複製):**")
            st.code(it["priv"], language="text")


# ==========================================
# 4. 側邊欄 (🛡️ 電腦/手機雙平台企業級資安防護版)
# ==========================================
doc, err = connect_google_sheets(sheet_url) if sheet_url else (None, "")
is_connected = bool(api_key and sheet_url and doc)

with st.sidebar:
    st.markdown("### ✦ BearJoy 導航")
    menu = st.radio("功能選單", ["智能客服系統", "問題分類庫", "折價券管理"], label_visibility="collapsed")
    st.markdown('<div style="flex-grow: 1; min-height: 50vh;"></div>', unsafe_allow_html=True)
    
    # 🛡️ 資安魔法：只要成功連線，不管手機或電腦，密碼框完全消失！
    if is_connected:
        st.markdown("""
        <div style='background-color:#E3DFD5; padding:12px; border-radius:8px; text-align:center; border: 1px solid #D0CCC1;'>
            <p style='color:#4A4238; font-weight:bold; font-size:15px; margin:0; margin-bottom:5px;'>✅ 系統已安全連線</p>
            <p style='color:#798571; font-size:12px; margin:0;'>金鑰已啟動最高級別隱藏防護</p>
        </div>
        """, unsafe_allow_html=True)
    else:
        # 只有在沒有金鑰或連線失敗時，才會顯示輸入框
        with st.expander("⚙️ 設定 (未偵測到有效金鑰)", expanded=True):
            input_api = st.text_input("API 金鑰:", value=api_key, type="password")
            input_url = st.text_input("試算表網址:", value=sheet_url)
            if st.button("儲存連線", use_container_width=True):
                save_config(input_api, input_url)
                st.rerun()
                
        if err:
            st.markdown(f"<p style='color:#A94442; font-weight:bold; font-size:14px; text-align:center;'>⚠️ 連線失敗: {err}</p>", unsafe_allow_html=True)

    # 💡 使用小提醒：休眠與「保持清醒」說明（給未來的自己看）
    st.markdown("<div style='margin-top:12px;'></div>", unsafe_allow_html=True)
    with st.expander("💡 開啟太慢/休眠畫面?"):
        # 字體大小與「✅ 系統已安全連線」一致(15px)，文字上下左右置中於白色框中
        st.markdown("""
        <div style='font-size:15px; color:#4A4238; line-height:1.75; text-align:left;'>
            <p style='margin:0 0 12px 0;'><b>為什麼要等一下?</b><br>
            免費雲端超過約 7 天沒人開會自動休眠。再開時按「喚醒」等 30 秒～1 分鐘即可,屬正常現象、不是當機。</p>
            <p style='margin:0 0 12px 0;'><b>想每次秒開</b><br>
            到 cron-job.org 把「保持 BearJoy 客服清醒」開關切 ON,定時戳網址讓系統不睡;不常用再切 OFF。</p>
            <p style='margin:0 0 12px 0;'><b>正確開啟順序</b><br>
            1. 先用手機開本頁、按「喚醒」<br>2. 再去 cron-job.org 切 ON<br>3. pinger 只能維持清醒、叫不醒睡著的</p>
            <p style='margin:0 0 12px 0;'><b>小提醒</b><br>
            左邊方框是「選取框」不是開關。開關請點 EDIT → Enabled 切換後存檔。</p>
            <p style='margin:0;'><b>建議間隔:每 6 小時或每天 1 次就夠,又省又不休眠。</b></p>
        </div>
        """, unsafe_allow_html=True)

# ==========================================
# 5. 主功能區
# ==========================================
if doc:
    if menu == "智能客服系統":
        st.markdown("""
        <div class="main-title-box">
            <div class="main-title-text">✦ BearJoy 智能客服系統 ✦</div>
        </div>
        """, unsafe_allow_html=True)
        
        tab1, tab2, tab3, tab5 = st.tabs(["批次評價處理", "VIP 顧客管理", "好評洞察 / 素材", "私訊查詢"])

        with tab1:
            col_up, col_res = st.columns([1, 1.5], gap="large")
            with col_up:
                # 手機版：上傳區與結果區改成上下單欄，框各自吃滿整個螢幕寬（電腦版維持左右並排）
                st.markdown('<span class="main-stack" style="display:none"></span>', unsafe_allow_html=True)
                st.markdown("##### ① 上傳好評截圖")
                files = st.file_uploader("上傳顧客好評截圖", type=["png", "jpg", "jpeg"], accept_multiple_files=True, label_visibility="collapsed")

                st.markdown("##### ② 回覆設定")
                cck1, cck2 = st.columns(2, vertical_alignment="center")
                cck1.markdown('<span class="cck-row" style="display:none"></span>', unsafe_allow_html=True)
                is_vip_check = cck1.checkbox("🌟 回購語氣")
                save_screenshots = cck2.checkbox("💾 保存截圖", value=True,
                                                 help="會把你上傳的截圖存到雲端「評價截圖素材」工作表，之後做素材用。會多花一點同步時間。")
                save_to_drive = st.checkbox("☁️ 同時備份原圖到 Google Drive 資料夾", value=True,
                                            help="把上傳的評價原圖存進你的 Drive 資料夾（檔名＝「日期 評價圖-規格」），"
                                                 "未來開資料夾就能直接挑素材。透過 Apps Script 以你本人身分存，會多花一點同步時間。")
                # 🎁 功能1：回購優惠碼設定收進摺疊區，平時不佔版面、要用再展開
                if "saved_repurchase" not in st.session_state:
                    try:
                        cfg_rows = get_or_create_ws(doc, "系統設定").get_all_values()
                        rc = next((r for r in cfg_rows if r and r[0] == "repurchase_code"), None)
                        ro = next((r for r in cfg_rows if r and r[0] == "repurchase_offer"), None)
                        st.session_state.saved_repurchase = (rc[1] if rc and len(rc) > 1 else "",
                                                             ro[1] if ro and len(ro) > 1 else "")
                    except Exception:
                        st.session_state.saved_repurchase = ("", "")
                _saved_code, _saved_offer = st.session_state.saved_repurchase
                with st.expander("🎁 回購優惠碼（選填）", expanded=bool(_saved_code)):
                    repurchase_code = st.text_input("回購優惠碼", value=_saved_code, placeholder="例如 BEARJOY50")
                    repurchase_offer = st.text_input("優惠說明", value=_saved_offer, placeholder="例如 全館滿299折20")
                    if st.button("💾 設為預設範例", use_container_width=True):
                        try:
                            cfg_ws = get_or_create_ws(doc, "系統設定")
                            _save_kv(cfg_ws, "repurchase_code", repurchase_code.strip())
                            _save_kv(cfg_ws, "repurchase_offer", repurchase_offer.strip())
                            st.session_state.saved_repurchase = (repurchase_code.strip(), repurchase_offer.strip())
                            st.success("已儲存為預設，下次打開會自動帶入 ✅")
                        except Exception as e:
                            st.error(f"儲存失敗：{e}")

                # ✍️ 這批要加強什麼：會一起交給 AI，套進每一則賣場回覆與私訊回覆
                batch_ins = st.text_input(
                    "✍️ 這批想加強／調整什麼？（選填）", key="batch_ins",
                    placeholder="例如：多提一句夏天出遊很適合、語氣再親切一點、提醒可回購")
                st.caption("留空就照原本語氣寫。跑完之後，每一筆還可以單獨再請 AI 加強。")

                st.markdown("##### ③ 選擇 AI 模型")
                _bkeys = list(GEMINI_PRICES.keys())
                batch_model = st.selectbox(
                    "AI 模型", _bkeys, index=_bkeys.index(GEMINI_DEFAULT),
                    format_func=lambda k: GEMINI_PRICES[k]["label"], key="batch_model",
                    label_visibility="collapsed",
                    help="預設為免費額度的推薦款；標「付費」的模型才會真的扣費。")

                st.markdown("##### ④ 開始處理")
                start_btn = st.button("🚀 開始解析並同步", type="primary", use_container_width=True)
                preview_area = st.container()

            with col_res:
                top_success_msg = st.empty()
                cards_container = st.container()
                
                # ▶️ 按下「開始處理」：整批交給背景執行緒，畫面只顯示進度
                #    （這樣手機把畫面滑掉、螢幕暗掉、甚至換頁，AI 都會繼續寫完。）
                if start_btn:
                    _running = (batch_job_get() or {}).get("status") == "running"
                    if _running:
                        st.warning("已經有一批正在背景處理中，等它跑完再開新的一批。")
                    elif not files:
                        st.warning("請先在①上傳好評截圖。")
                    elif not api_key:
                        st.warning("尚未設定 API 金鑰。")
                    else:
                        # 圖片要在這裡就整份讀進記憶體：背景執行緒讀不到瀏覽器上傳的檔案物件
                        _imgs = []
                        for _f in files:
                            try:
                                _im = Image.open(_f)
                                _im.load()
                                _imgs.append((_f.name, _im.convert("RGB")))
                            except Exception as _e:
                                st.caption(f"⚠️ {_f.name} 讀不進來，已略過：{_e}")
                        _creds = None
                        try:
                            if "type" in st.secrets:
                                _creds = dict(st.secrets)
                        except Exception:
                            pass
                        if _imgs:
                            st.session_state.pop("batch_results", None)
                            st.session_state.pop("batch_msg", None)
                            batch_job_start(_imgs, {
                                "api_key": api_key,
                                "model": st.session_state.get("batch_model", GEMINI_DEFAULT),
                                "is_vip": is_vip_check,
                                "batch_ins": (batch_ins or "").strip(),
                                "repurchase_code": repurchase_code.strip(),
                                "repurchase_offer": repurchase_offer.strip(),
                                "save_screenshots": save_screenshots,
                                "save_to_drive": save_to_drive,
                                "creds_dict": _creds,
                                "sheet_url": sheet_url,
                            })
                            st.rerun()

                # 🏃 背景進度區：每 2 秒自己更新一次；跑完自動換成下方結果區
                _job = batch_job_get()
                if _job and _job.get("status") == "running":
                    render_batch_progress()
                elif _job and _job.get("status") in ("done", "error"):
                    if not _job.get("_merged"):
                        _job["_merged"] = True
                        st.session_state["batch_results"] = [
                            {"acc": it["acc"], "rev": it["rev"], "pub": it["pub"],
                             "priv": it["priv"], "row": it.get("row")}
                            for it in _job.get("items", [])]
                        if _job.get("msg"):
                            st.session_state["batch_msg"] = _job["msg"]
                        if _job.get("cost"):
                            st.session_state["ai_cost_total"] = \
                                st.session_state.get("ai_cost_total", 0.0) + _job["cost"]
                        if _job.get("last_usage"):
                            st.session_state["ai_last_usage"] = _job["last_usage"]
                    with cards_container:
                        _fin = _job.get("finished", "")
                        st.caption(f"🏁 這批 {_job.get('total', 0)} 筆已跑完"
                                   f"（{_job.get('started','')} → {_fin}），共成功 "
                                   f"{len(_job.get('items', []))} 筆。")
                        if _job.get("drive_ok") or _job.get("drive_fail"):
                            _dm = (f"☁️ Drive 備份：成功 {_job.get('drive_ok', 0)} 筆"
                                   f"／失敗 {_job.get('drive_fail', 0)} 筆")
                            (st.warning if _job.get("drive_fail") else st.caption)(_dm)
                        if _job.get("status") == "error":
                            st.error(f"這批中途出錯：{_job.get('error', '')}")
                        for _n in _job.get("notes", []):
                            st.caption(_n)

                # 📋 上一批的結果（重跑也不會不見）：可逐筆下指示請 AI 改寫，改完同步寫回雲端那一列
                _bres = st.session_state.get("batch_results") or []
                if _bres and not start_btn:
                    if st.session_state.get("batch_msg"):
                        top_success_msg.success(st.session_state["batch_msg"])

                    def _batch_rewrite(i, field, col_no):
                        """依加強指示改寫第 i 筆的賣場回覆(field='pub') 或私訊回覆(field='priv')，
                        成功就更新畫面內容，並寫回雲端「回覆紀錄」同一列（寫前先核對帳號，避免寫錯列）。"""
                        it = st.session_state["batch_results"][i]
                        ins = (st.session_state.get(f"br_ins_{i}", "") or "").strip()
                        if not api_key:
                            st.warning("尚未設定 API 金鑰。")
                            return
                        if not ins:
                            st.warning("請先在「想加強／調整什麼」寫一句話，AI 才知道要往哪改。")
                            return
                        with st.spinner("AI 改寫中…"):
                            out, usage = qa_ai_refine(
                                api_key, it[field], ins, category="顧客評價回覆", question=it["rev"],
                                model=st.session_state.get("batch_model", GEMINI_DEFAULT))
                        ai_track_cost(usage)
                        if not out:
                            st.error("AI 改寫失敗，請稍後再試。")
                            return
                        st.session_state["batch_results"][i][field] = out
                        if doc and it.get("row"):
                            try:
                                _ws = get_or_create_ws(doc, "回覆紀錄")
                                if str(_ws.cell(it["row"], 2).value or "").strip() == str(it["acc"]).strip():
                                    _ws.update_cell(it["row"], col_no, out)
                                else:
                                    st.caption("⚠️ 雲端那一列對不上帳號，這次只改畫面沒有寫回雲端。")
                            except Exception as e:
                                st.caption(f"⚠️ 寫回雲端失敗（畫面已更新）：{e}")
                        st.rerun()

                    def _batch_fix_account(i):
                        """✏️ 當場更正 AI 認錯的客戶帳號：同步改雲端「回覆紀錄」那一列的帳號，
                        並把 VIP 名單的互動次數從錯帳號搬到正確帳號（錯的那筆歸零就刪掉）。"""
                        it = st.session_state["batch_results"][i]
                        old_acc = str(it.get("acc", "")).strip()
                        new_acc = (st.session_state.get(f"br_acc_{i}", "") or "").strip()
                        if not new_acc:
                            st.warning("請先填正確的客戶帳號。")
                            return
                        if new_acc == old_acc:
                            st.info("帳號沒有變更。")
                            return
                        st.session_state["batch_results"][i]["acc"] = new_acc
                        if not (doc and it.get("row")):
                            st.caption("⚠️ 這筆沒有對應的雲端列號，只改了畫面。")
                            st.rerun()
                        try:
                            _ws = get_or_create_ws(doc, "回覆紀錄")
                            if str(_ws.cell(it["row"], 2).value or "").strip() != old_acc:
                                st.caption("⚠️ 雲端那一列對不上原帳號，這次只改畫面沒有寫回雲端。")
                                st.rerun()
                            _ws.update_cell(it["row"], 2, new_acc)
                            # 👑 VIP 名單跟著搬：錯帳號 -1（歸零就刪列），正確帳號 +1（沒有就新增）
                            try:
                                _wv = get_or_create_ws(doc, "VIP名單")
                                _vals = _wv.get_all_values()
                                _today = datetime.now().strftime("%Y-%m-%d")
                                _idx_old = next((r for r, v in enumerate(_vals[1:], start=2)
                                                 if v and str(v[0]).strip() == old_acc), None)
                                if _idx_old:
                                    try:
                                        _cnt = int(str(_vals[_idx_old - 1][3]).strip() or 1)
                                    except Exception:
                                        _cnt = 1
                                    if _cnt <= 1:
                                        _wv.delete_rows(_idx_old)
                                        _vals = _wv.get_all_values()
                                    else:
                                        _wv.update_cell(_idx_old, 4, _cnt - 1)
                                _idx_new = next((r for r, v in enumerate(_vals[1:], start=2)
                                                 if v and str(v[0]).strip() == new_acc), None)
                                if _idx_new:
                                    try:
                                        _cn = int(str(_vals[_idx_new - 1][3]).strip() or 0)
                                    except Exception:
                                        _cn = 0
                                    _wv.update_cell(_idx_new, 3, _today)
                                    _wv.update_cell(_idx_new, 4, _cn + 1)
                                else:
                                    _wv.append_row([new_acc, _today, _today, 1])
                            except Exception as e:
                                st.caption(f"⚠️ 回覆紀錄已更正，但 VIP 名單沒改到：{str(e)[:80]}")
                            st.success(f"已把「{old_acc}」更正為「{new_acc}」，雲端同步完成 ✅")
                        except Exception as e:
                            st.caption(f"⚠️ 寫回雲端失敗（畫面已更新）：{e}")
                        st.rerun()

                    with cards_container:
                        for _i, _it in enumerate(_bres):
                            with st.expander(f"✨ 客戶帳號：{_it['acc']}", expanded=True):
                                # ✏️ AI 有時會把帳號看錯：這裡可以當場改，按下去就同步雲端
                                _ac1, _ac2 = st.columns([2, 1], vertical_alignment="bottom")
                                _ac1.markdown('<span class="keep-row" style="display:none"></span>',
                                              unsafe_allow_html=True)
                                _ac1.text_input("👤 客戶帳號（AI 認錯可直接改）", value=_it["acc"],
                                                key=f"br_acc_{_i}")
                                if _ac2.button("💾 更正帳號", key=f"br_accfix_{_i}",
                                               use_container_width=True):
                                    _batch_fix_account(_i)
                                st.markdown(f"**📝 原始評價內容:** {_it['rev']}")
                                st.markdown("**📢 賣場回覆 (點擊右上角複製):**")
                                st.code(_it["pub"], language="text")
                                st.markdown("**💌 私訊回覆 (點擊右上角複製):**")
                                st.code(_it["priv"], language="text")
                                st.text_input("✍️ 想加強／調整什麼？（AI 會融進整則回覆）",
                                              key=f"br_ins_{_i}",
                                              placeholder="例如：多謝謝他提到的材質、語氣再親切一點")
                                _q1, _q2 = st.columns(2)
                                if _q1.button("🤖 改寫賣場回覆", key=f"br_pub_{_i}", use_container_width=True):
                                    _batch_rewrite(_i, "pub", 4)
                                if _q2.button("🤖 改寫私訊回覆", key=f"br_priv_{_i}", use_container_width=True):
                                    _batch_rewrite(_i, "priv", 5)
                        ai_render_cost(st.session_state.get("batch_model", GEMINI_DEFAULT))
                        if st.button("🧹 清掉這批結果", key="br_clear", use_container_width=True):
                            for _k in [k for k in list(st.session_state.keys())
                                       if str(k).startswith(("br_ins_", "br_acc_"))]:
                                st.session_state.pop(_k, None)
                            st.session_state.pop("batch_results", None)
                            st.session_state.pop("batch_msg", None)
                            batch_job_store()["job"] = None   # 連背景任務紀錄一起清掉
                            st.rerun()

        with tab2:
            with st.container(border=True):
                section_block("👑", "VIP 顧客戰情室", "所有與你互動過的顧客名單與互動次數。最多顯示 5 筆，其餘用表格右側滾輪查看。")
                if doc:
                    try:
                        vip_ws = get_or_create_ws(doc, "VIP名單")
                        data = vip_ws.get_all_values()
                        if len(data) > 1:
                            st.caption(f"目前共 {len(data) - 1} 位 VIP 顧客")
                            # ✨ 固定高度＝表頭＋5 列：最多呈現 5 筆，其餘在框內用右側滾輪捲動，不會把整頁撐長
                            _vipdf = pd.DataFrame(data[1:], columns=data[0])
                            # 依「最後互動」日期由新到舊排序（最新的排最上面）
                            if "最後互動" in _vipdf.columns:
                                _vipdf = (_vipdf.assign(_k=pd.to_datetime(_vipdf["最後互動"], errors="coerce"))
                                                .sort_values("_k", ascending=False, na_position="last")
                                                .drop(columns="_k").reset_index(drop=True))
                            st.dataframe(_vipdf, use_container_width=True, height=213)
                        else: st.info("目前 VIP 名單尚無資料，趕快去解析第一筆評價吧！")

                        # 💤 功能3：沉睡客喚回——找出好久沒回來的老客，一鍵生成喚回訊息
                        if len(data) > 1:
                            section_block("💤", "沉睡客喚回", "找出好久沒回來的老客，生成專屬喚回訊息＋優惠碼，貼到蝦皮聊聊就能發。建議 30～90 天；想測試可先把天數設小一點看效果。")
                            c_days, c_code = st.columns(2)
                            c_days.markdown('<span class="keep-row" style="display:none"></span>', unsafe_allow_html=True)
                            days = c_days.number_input("幾天沒互動就算沉睡客?", min_value=1, max_value=365, value=30, step=1, key="sleep_days")
                            wb_code = c_code.text_input("喚回專屬優惠碼（選填）", placeholder="例如 COMEBACK50", key="wb_code")
                            header = data[0]
                            i_acc = header.index("客戶帳號") if "客戶帳號" in header else 0
                            i_last = header.index("最後互動") if "最後互動" in header else 2
                            today = datetime.now()
                            sleepers, parsed = [], 0
                            for r in data[1:]:
                                if len(r) <= max(i_acc, i_last):
                                    continue
                                last = str(r[i_last]).strip()[:10]
                                try:
                                    gap = (today - datetime.strptime(last, "%Y-%m-%d")).days
                                    parsed += 1
                                    if gap >= int(days):
                                        sleepers.append((str(r[i_acc]), last, gap))
                                except Exception:
                                    continue
                            if sleepers:
                                sleepers.sort(key=lambda x: -x[2])
                                st.write(f"共找到 **{len(sleepers)}** 位沉睡客（依沉睡天數排序）：")
                                coupon_line = f"為感謝您的支持，送上專屬優惠碼 👉 {wb_code.strip()} 🎁\n" if wb_code.strip() else ""

                                def _wakeback_msg(acc):
                                    return (f"親愛的 {acc}，\n\n"
                                            f"好久不見了，BearJoy Sharon 一直記得您 🥰\n"
                                            f"最近上了新品與優惠，特別想第一個和您分享！\n"
                                            f"{coupon_line}"
                                            f"期待您再回來逛逛 ❤️\n\n—— BearJoy Sharon")

                                # 📥 下載沉睡客名單 Excel（含帳號、最後互動、沉睡天數、可直接複製的喚回訊息）
                                try:
                                    df_sleep = pd.DataFrame(
                                        [{"客戶帳號": acc, "最後互動日期": last, "已沉睡天數": gap,
                                          "喚回訊息": _wakeback_msg(acc)} for acc, last, gap in sleepers])
                                    xbuf = BytesIO()
                                    with pd.ExcelWriter(xbuf, engine="openpyxl") as writer:
                                        df_sleep.to_excel(writer, index=False, sheet_name="沉睡客名單")
                                        _excel_align_left_top(writer.sheets["沉睡客名單"])
                                    cdl1, cdl2 = st.columns(2)
                                    cdl1.markdown('<span class="keep-row" style="display:none"></span>', unsafe_allow_html=True)
                                    with cdl1:
                                        st.download_button(
                                            "📥 下載 Excel",
                                            data=xbuf.getvalue(),
                                            file_name=f"BearJoy_沉睡客名單_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                            use_container_width=True)
                                    with cdl2:
                                        if st.button("☁️ 同步到雲端", use_container_width=True, key="sleep_cloud"):
                                            try:
                                                with st.spinner("寫入雲端中…"):
                                                    link = write_df_to_sheet(doc, "沉睡客名單", df_sleep)
                                                st.session_state.sleep_sync_msg = ("ok", "已寫入雲端 Google Sheet「沉睡客名單」分頁 ✅", link)
                                            except Exception as e:
                                                st.session_state.sleep_sync_msg = ("err", f"雲端同步失敗：{e}", None)
                                    # 同步結果訊息：滿版顯示在兩顆按鍵下方，不擠動上面的「下載 Excel」鍵
                                    _slp_msg = st.session_state.get("sleep_sync_msg")
                                    if _slp_msg:
                                        _mk, _mtext, _mlink = _slp_msg
                                        if _mk == "ok":
                                            st.success(_mtext)
                                            if _mlink:
                                                st.markdown(f"[👉 點此開啟雲端名單]({_mlink})")
                                        else:
                                            st.error(_mtext)
                                except Exception as e:
                                    st.caption(f"⚠️ Excel 名單產生失敗（不影響下方訊息）：{e}")

                                for acc, last, gap in sleepers:
                                    with st.expander(f"💤 {acc}（已 {gap} 天沒互動，最後 {last}）"):
                                        st.code(_wakeback_msg(acc), language="text")
                            elif parsed == 0:
                                st.warning("讀不到「最後互動」日期，請確認 VIP名單 的日期格式為 2026-06-02。")
                            else:
                                st.success(f"已檢查 {parsed} 位顧客，目前沒有超過 {int(days)} 天沒互動的沉睡客 🎉")
                    except Exception as e:
                        st.error(f"讀取失敗：{e}")

        with tab3:
            # 共用：載入文字評價清單（最新在前）；分析與好評圖挑選都用這份，避免重複讀取
            if "review_pool" not in st.session_state or st.session_state.get("refresh_review_pool"):
                try:
                    _hist = get_or_create_ws(doc, "回覆紀錄").get_all_values()
                    _rows = [r for r in _hist[1:] if len(r) > 2 and r[2].strip() and r[2].strip() != "解析失敗"]
                    _rows = list(reversed(_rows))  # 最新在前
                    # 🔁 去重：同一帳號＋同一評價內容只保留一筆(最新)，避免重複處理時同一則好評出現兩次
                    _seen_rv = set(); _dedup_rv = []
                    for _r in _rows:
                        _acc = _r[1].strip() if len(_r) > 1 else ""
                        _key = (_acc, _r[2].strip())
                        if _key in _seen_rv:
                            continue
                        _seen_rv.add(_key); _dedup_rv.append(_r)
                    st.session_state.review_pool = _dedup_rv
                except Exception:
                    st.session_state.review_pool = []
                st.session_state.refresh_review_pool = False
            review_pool = st.session_state.review_pool

            # 📊 大區塊一：顧客最愛優點分析
            with st.container(border=True):
                section_block("📊", "顧客最愛優點分析", "選規格 → 統整該款顧客最愛優點，直接拿去寫蝦皮標題與賣點。")
                _specs_all = sorted({_review_spec(r[2]) for r in review_pool if _review_spec(r[2])})
                fic1, fic2 = st.columns(2)
                sel_specs = fic1.multiselect("選規格（可複選＝合併；可打字搜尋；留空＝全部）",
                                             _specs_all, key="insight_specs")
                kw = fic2.text_input("或用關鍵字一次分析", placeholder="例：三層 → 所有含三層的款",
                                     key="insight_kw").strip()
                if st.button("🔍 開始分析", use_container_width=True):
                    if not api_key:
                        st.error("需要 API 金鑰才能分析。")
                    else:
                        try:
                            if sel_specs:
                                reviews = [r[2] for r in review_pool if _review_spec(r[2]) in sel_specs]
                                used_label = "、".join(sel_specs)
                            elif kw:
                                reviews = [r[2] for r in review_pool if kw.lower() in _review_spec(r[2]).lower() and _review_spec(r[2])]
                                used_label = f"關鍵字「{kw}」"
                            else:
                                reviews = [r[2] for r in review_pool]
                                used_label = "全部"
                            if not reviews:
                                st.info("找不到符合的評價可分析（換個關鍵字或規格試試）。")
                            else:
                                with st.spinner(f"AI 正在分析 {len(reviews)} 筆評價..."):
                                    joined = "\n".join(f"- {r}" for r in reviews[:200])
                                    prompt = (
                                        "以下是某蝦皮賣場的真實顧客評價。請統整出顧客最常稱讚的「5 個優點」，"
                                        "用 Markdown 條列，每點格式為：\n"
                                        "**1. 優點標題** — 簡短說明（附 1 句顧客原話佐證）\n"
                                        "最後另起一段，以 **🛒 賣點建議：** 開頭，給 2~3 句可直接用在商品標題或描述的文案。\n\n"
                                        f"顧客評價如下：\n{joined}"
                                    )
                                    result = gemini_generate(api_key, [prompt])
                                if result:
                                    st.session_state.insight_result = result
                                    st.session_state.insight_spec_used = used_label
                                    st.session_state.pop("insight_sync_msg", None)  # 新分析→清掉上次的同步訊息
                                else:
                                    st.error("分析失敗，請稍後再試（可能是 AI 額度或網路問題）。")
                        except Exception as e:
                            st.error(f"分析失敗：{e}")
                if st.session_state.get("insight_result"):
                    _used = st.session_state.get("insight_spec_used", "全部")
                    st.markdown(f"<p style='font-size:12px;color:#A39C90;margin:6px 0 2px 2px;'>分析範圍：{_used}</p>", unsafe_allow_html=True)
                    with st.container(border=True):
                        st.markdown(st.session_state.insight_result)
                    # 第一欄＝規格（哪一款），分析內容整段放同一格
                    df_insight = pd.DataFrame([{
                        "規格": st.session_state.get("insight_spec_used", "全部"),
                        "顧客優點分析": _strip_md(st.session_state.insight_result),
                    }])
                    cin1, cin2 = st.columns(2)
                    cin1.markdown('<span class="keep-row" style="display:none"></span>', unsafe_allow_html=True)
                    with cin1:
                        try:
                            ibuf = BytesIO()
                            with pd.ExcelWriter(ibuf, engine="openpyxl") as writer:
                                df_insight.to_excel(writer, index=False, sheet_name="顧客優點分析")
                                _excel_align_left_top(writer.sheets["顧客優點分析"])
                            st.download_button(
                                "📥 下載 Excel",
                                data=ibuf.getvalue(),
                                file_name=f"BearJoy_顧客優點分析_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True)
                        except Exception as e:
                            st.caption(f"⚠️ Excel 產生失敗：{e}")
                    with cin2:
                        if st.button("☁️ 同步到雲端", use_container_width=True, key="insight_cloud"):
                            try:
                                with st.spinner("寫入雲端中…"):
                                    link = write_df_to_sheet(doc, "顧客優點分析", df_insight)
                                st.session_state.insight_sync_msg = ("ok", "已寫入雲端 Google Sheet「顧客優點分析」分頁 ✅", link)
                            except Exception as e:
                                st.session_state.insight_sync_msg = ("err", f"雲端同步失敗：{e}", None)
                    # 同步結果訊息：滿版顯示在兩顆按鍵下方（從下載 Excel 鍵到同步雲端鍵下方都可呈現），不擠動上面的按鍵
                    _ins_msg = st.session_state.get("insight_sync_msg")
                    if _ins_msg:
                        _mk, _mtext, _mlink = _ins_msg
                        if _mk == "ok":
                            st.success(_mtext)
                            if _mlink:
                                st.markdown(f"[👉 點此開啟雲端分析]({_mlink})")
                        else:
                            st.error(_mtext)

            # 🖼️ 大區塊二：一鍵生成顧客好評圖
            with st.container(border=True):
                section_block("🖼️", "一鍵生成「顧客好評圖」", "貼蝦皮置頂、IG、FB、LINE，提升新客下單信任感。")
                SIZE_PRESETS = {
                    "正方形 1:1（IG / 蝦皮 1080×1080）": (1080, 1080),
                    "直式 9:16（限動・Reels 1080×1920）": (1080, 1920),
                    "橫式（FB貼文 1200×630）": (1200, 630),
                    "LINE 圖文（1040×1040）": (1040, 1040),
                    "自訂尺寸…": None,
                }
                # 版型 + 自動取幾筆：同一排（手機版維持並排；幾筆較窄）；圖片尺寸自己一排
                c_tpl, c_n = st.columns([2, 1])
                c_tpl.markdown('<span class="ratio-row" style="display:none"></span>', unsafe_allow_html=True)
                template_label = c_tpl.selectbox("版型", [
                    "版型A：真實截圖拼接",
                    "版型B：文字精選卡",
                    "版型C：大字引用感",
                ], key="rev_tpl")
                rev_n = c_n.number_input("或 取幾筆", min_value=1, max_value=12, value=3, step=1,
                                         key="rev_img_n", help="與『手動勾選評價』二擇一：沒手動勾選時，才會自動取最新這幾筆")
                size_label = st.selectbox("圖片尺寸", list(SIZE_PRESETS.keys()), key="rev_size")
                target_size = SIZE_PRESETS[size_label]
                if target_size is None:
                    cw, ch = st.columns(2)
                    cust_w = cw.number_input("寬 (px)", 300, 4000, 1080, 20, key="rev_cw")
                    cust_h = ch.number_input("高 (px)", 300, 4000, 1080, 20, key="rev_ch")
                    target_size = (int(cust_w), int(cust_h))

                # 只顯示「當前版型」對應的挑選區，畫面更乾淨
                mat_pool = []
                pool_for_img = review_pool  # 版型B/C 做圖用的評價範圍，下方規格/關鍵字篩選會覆蓋它
                if template_label.startswith("版型A"):
                    if "mat_pool" not in st.session_state or st.session_state.get("refresh_mat_pool"):
                        try:
                            _mat = get_or_create_ws(doc, "評價截圖素材").get_all_values()
                            _mat = [r for r in _mat if len(r) > 1 and r[0]]
                            st.session_state.mat_pool = list(reversed(_mat))[:40]
                        except Exception:
                            st.session_state.mat_pool = []
                        st.session_state.refresh_mat_pool = False
                    mat_pool = st.session_state.mat_pool
                    # 帳號→規格 對照（給舊截圖沒存規格時補規格）
                    acc2spec = {}
                    for rr in review_pool:
                        sp = _review_spec(rr[2])
                        if sp and rr[1] not in acc2spec:
                            acc2spec[rr[1]] = sp
                    def _mat_spec(r):
                        _, a, sp = _mat_meta(r[0])
                        return _spec_no_color(sp) if sp else acc2spec.get(a, "")
                    # 🔎 規格/關鍵字篩選（與「顧客最愛優點分析」一致）：選項用全部規格清單，4 個款都會出現
                    ma_specs = st.multiselect("選規格（可複選；留空＝全部）",
                                              _specs_all, key="matimg_specs")
                    ma_kw = st.text_input("或用關鍵字篩選", placeholder="例：三層", key="matimg_kw").strip()
                    # 🔍 AI 補抓規格：對「無規格」的舊截圖，用 AI 直接從圖片讀出規格並補進素材
                    _no_spec = [r for r in mat_pool if not _mat_spec(r)]
                    if _no_spec and st.button(f"🔍 AI 補抓規格（{len(_no_spec)} 張沒規格的截圖）", key="ai_fill_spec"):
                        if not api_key:
                            st.error("需要 API 金鑰才能補抓。")
                        else:
                            try:
                                ws_mat2 = get_or_create_ws(doc, "評價截圖素材")
                                idmap = {rr[0]: i + 1 for i, rr in enumerate(ws_mat2.get_all_values()) if rr and rr[0]}
                                done = 0
                                with st.spinner(f"AI 正在從圖片補抓 {len(_no_spec)} 張的規格…"):
                                    for r in _no_spec:
                                        try:
                                            im = base64_chunks_to_img([c for c in r[1:] if c])
                                            sp = (gemini_generate(api_key, ["只回答這張蝦皮評價截圖中顧客購買的『規格/分類/款式』，例如：[三層款 可拆卸 十合一]。要含中括號內款式，但『不要顏色』（不要黑色/粉色那種）。找不到就只回『無』。只回規格本身，不要任何多餘文字。", im]) or "").strip()
                                            sp = _spec_no_color(sp.splitlines()[0].strip()) if sp else ""
                                            if sp and sp != "無" and idmap.get(r[0]):
                                                base = r[0].split("|||")[0]
                                                ws_mat2.update_cell(idmap[r[0]], 1, f"{base}|||{sp}")
                                                done += 1
                                            time.sleep(2)
                                        except Exception:
                                            continue
                                st.session_state.refresh_mat_pool = True
                                st.success(f"已補抓 {done} 張的規格 ✅ 自動重新整理中…")
                                st.rerun()
                            except Exception as e:
                                st.error(f"補抓失敗：{e}")
                    def _mat_match(r):
                        sp = _mat_spec(r)
                        if ma_specs:
                            return sp in ma_specs
                        if ma_kw:
                            return bool(sp) and ma_kw.lower() in sp.lower()
                        return True
                    mat_filtered = [(i, r) for i, r in enumerate(mat_pool) if _mat_match(r)]
                    with st.expander("🖼️ 挑選要拼接的截圖（不挑＝用最新幾張）", expanded=True):
                        if not mat_filtered:
                            st.caption("這個規格／關鍵字下沒有截圖，換個條件；或先到「批次評價處理」勾『💾 保存截圖』。")
                        else:
                            st.caption(f"符合 {len(mat_filtered)} 張。勾選想要的；框內可上下捲動。")
                            thumbs = st.session_state.setdefault("_mat_thumbs", {})
                            with st.container(height=330):
                                cols = st.columns(3)
                                for n, (idx, r) in enumerate(mat_filtered):
                                    with cols[n % 3]:
                                        th = thumbs.get(r[0])
                                        if th is None:
                                            try:
                                                im = base64_chunks_to_img([c for c in r[1:] if c]); im.thumbnail((400, 400)); th = im
                                            except Exception:
                                                th = False
                                            thumbs[r[0]] = th
                                        if th:
                                            st.image(th, use_container_width=True)
                                        _sp = _mat_spec(r)
                                        st.checkbox(f"選 ［{_sp or '無規格'}］", key=f"matpick_{idx}")
                else:
                    # 🔎 規格/關鍵字篩選（與「顧客最愛優點分析」一致）：先縮小要做圖的評價範圍
                    img_specs = st.multiselect("選規格（可複選；留空＝全部）", _specs_all, key="revimg_specs")
                    img_kw = st.text_input("或用關鍵字篩選", placeholder="例：三層", key="revimg_kw").strip()
                    if img_specs:
                        pool_for_img = [r for r in review_pool if _review_spec(r[2]) in img_specs]
                    elif img_kw:
                        pool_for_img = [r for r in review_pool if img_kw.lower() in _review_spec(r[2]).lower() and _review_spec(r[2])]
                    else:
                        pool_for_img = review_pool
                    with st.expander("✋ 自己挑要放哪幾筆好評（每筆顯示完整內容）", expanded=False):
                        if pool_for_img:
                            st.caption(f"符合 {len(pool_for_img)} 筆。勾選想要的；框內可上下捲動。")
                            with st.container(height=360):
                                for i, r in enumerate(pool_for_img):
                                    spec = _review_spec(r[2])
                                    st.checkbox(f"#{i + 1}　{r[1]}　［{spec or '無規格'}］", key=f"revpick_{i}")
                                    _content = str(r[2]).replace("<", "&lt;").replace(">", "&gt;")
                                    st.markdown(
                                        f"<div style='font-size:13px;color:#4A4238;line-height:1.5;margin:-6px 0 10px 26px;'>{_content}</div>",
                                        unsafe_allow_html=True)
                        elif img_specs or img_kw:
                            st.caption("這個規格／關鍵字下沒有評價，換一個條件試試。")
                        else:
                            st.caption("目前沒有可挑選的文字評價，先去「批次評價處理」處理幾筆吧。")

                rcol, gcol, zcol = st.columns(3)
                rcol.markdown('<span class="trio-btn" style="display:none"></span>', unsafe_allow_html=True)
                if rcol.button("🔄 重新整理", use_container_width=True, key="refresh_pool_all"):
                    st.session_state.refresh_mat_pool = True
                    st.session_state.refresh_review_pool = True
                    st.rerun()
                gen_clicked = gcol.button("✨ 產生好評圖", use_container_width=True)
                zip_clicked = zcol.button("📦 打包原圖", use_container_width=True)
                if gen_clicked:
                    try:
                        card = None
                        if template_label.startswith("版型A"):
                            # 版型A：用實際評價截圖拼接；有勾選就用勾的，沒勾就用最新幾張
                            sel_idx = [i for i, _ in mat_filtered if st.session_state.get(f"matpick_{i}")]
                            chosen = [mat_pool[i] for i in sel_idx] if sel_idx else [r for _, r in mat_filtered][:int(rev_n)]
                            imgs = []
                            for r in chosen:
                                try:
                                    imgs.append(base64_chunks_to_img([c for c in r[1:] if c]))
                                except Exception:
                                    continue
                            if not imgs:
                                st.info("還沒有已保存的評價截圖。請先到「批次評價處理」勾選『💾 保存截圖』並處理幾筆，再回來生成。")
                            else:
                                card = make_collage_image(imgs, size=target_size)
                        else:
                            # 版型B/C：用 AI 解析後的文字做圖
                            # 有勾選就用勾選的；沒勾就自動取最新、且「只取有規格」的（確保每則都有規格、也避免新舊重複）
                            picked = [i for i in range(len(pool_for_img)) if st.session_state.get(f"revpick_{i}")]
                            if picked:
                                rows = [pool_for_img[i] for i in picked]
                            else:
                                spec_rows = [r for r in pool_for_img if _review_spec(r[2])]
                                rows = (spec_rows or pool_for_img)[:int(rev_n)]
                            reviews = [(r[1], r[2]) for r in rows]
                            if not reviews:
                                st.info("還沒有評價可以做圖，先處理幾筆好評吧！")
                            else:
                                tpl = "quote" if template_label.startswith("版型C") else "cards"
                                card = make_review_image(reviews, size=target_size, template=tpl)
                        if card is not None:
                            st.image(card, use_container_width=True)
                            st.caption("💡 手機：長按上方圖片即可儲存，或按下方按鈕下載到手機。")
                            buf = BytesIO(); card.save(buf, format="PNG")
                            st.download_button("💻 下載好評圖", data=buf.getvalue(),
                                               file_name=f"BearJoy_好評圖_{target_size[0]}x{target_size[1]}.png", mime="image/png")
                    except Exception as e:
                        st.error(f"產生失敗：{e}")

                # 📥 打包下載評價原圖（按鈕在上方與「產生好評圖」同排，給美編用）
                if zip_clicked:
                    try:
                        import zipfile
                        mat = get_or_create_ws(doc, "評價截圖素材").get_all_values()
                        mat = [r for r in mat if len(r) > 1 and r[0]]
                        if not mat:
                            st.info("還沒有已保存的評價原圖。請先到「批次評價處理」勾『💾 保存截圖』並處理幾筆。")
                        else:
                            # 帳號 → 規格 對照（從回覆紀錄取，用來命名）
                            acc2spec = {}
                            for r in review_pool:
                                sp = _review_spec(r[2])
                                if sp and r[1] not in acc2spec:
                                    acc2spec[r[1]] = sp
                            zbuf = BytesIO()
                            used = {}
                            with zipfile.ZipFile(zbuf, "w", zipfile.ZIP_DEFLATED) as zf:
                                for r in mat:
                                    try:
                                        im = base64_chunks_to_img([c for c in r[1:] if c])
                                    except Exception:
                                        continue
                                    date, acc, spec = _mat_meta(r[0])
                                    base = f"{date} 評價圖-{_safe_filename(_spec_no_color(spec) if spec else (acc2spec.get(acc) or acc))}"
                                    k = used.get(base, 0); used[base] = k + 1
                                    fname = f"{base}.png" if k == 0 else f"{base}_{k + 1}.png"
                                    ib = BytesIO(); im.convert("RGB").save(ib, format="PNG")
                                    zf.writestr(fname, ib.getvalue())
                            st.success(f"已打包 {sum(used.values())} 張原圖 ✅")
                            st.download_button("💾 下載 ZIP", data=zbuf.getvalue(),
                                               file_name=f"BearJoy_評價原圖_{datetime.now().strftime('%Y%m%d')}.zip",
                                               mime="application/zip")
                    except Exception as e:
                        st.error(f"打包失敗：{e}")

        # ==========================================
        # 💌 Tab5：私訊查詢（搜尋客戶 → 看當初那則評價 → 複製要私訊的內容；不做狀態管理）
        # ==========================================
        with tab5:
            with st.container(border=True):
                section_block("💌", "私訊查詢",
                              "公開回覆先發、私訊過幾天再回時用這裡：打客戶帳號（或評價裡的任何字）就會叫出"
                              "當初那則評價，以及當時幫你寫好的私訊內容，點右上角圖示複製直接貼給客人。")
                if not doc:
                    st.info("請先在左側完成連線，才能查詢。")
                else:
                    try:
                        dm_ws, dm_rows = dm_load(doc)
                    except Exception as e:
                        dm_ws, dm_rows = None, []
                        st.error(f"讀取回覆紀錄失敗：{e}")

                    if dm_ws is not None:
                        dm_kw = st.text_input("搜尋", key="dm_kw", label_visibility="collapsed",
                                              placeholder="🔍 打客戶帳號或評價裡的字（例如 hasqwsky、三層款）")

                        def _dm_match(r):
                            kws = (dm_kw or "").lower().split()
                            if not kws:
                                return True
                            blob = " ".join([str(r.get(c, "")) for c in
                                             ("客戶帳號", "原始評價內容", "VIP私訊回覆", "紀錄時間")]).lower()
                            return all(t in blob for t in kws)

                        _searching = bool((dm_kw or "").strip())
                        shown_dm = [r for r in dm_rows if _dm_match(r)]
                        _limit = len(shown_dm) if _searching else 8

                        if not dm_rows:
                            st.info("還沒有任何紀錄。到「批次評價處理」跑過一批評價之後，私訊內容就會出現在這裡。")
                        elif _searching:
                            st.caption(f"找到 {len(shown_dm)} 筆（共 {len(dm_rows)} 筆紀錄）")
                            if not shown_dm:
                                st.info("找不到符合的紀錄，換個關鍵字試試（帳號大小寫沒關係）。")
                        else:
                            st.caption(f"共 {len(dm_rows)} 筆紀錄，先顯示最新 8 筆；要找特定客人請在上面搜尋。")

                        for r in shown_dm[:_limit]:
                            with st.container(border=True):
                                _acc = (r.get("客戶帳號", "") or "未知").strip()
                                _t = (r.get("紀錄時間", "") or "").strip()
                                st.markdown(f"#### 👤 {_acc}")
                                st.caption(f"🕒 {_t}")
                                _rev = (r.get("原始評價內容", "") or "").strip()
                                if _rev:
                                    st.markdown("**📝 當初的評價：**")
                                    st.markdown(
                                        "<div style='background:#FFFFFF; border:1px solid #E6E2D8; border-radius:8px;"
                                        "padding:10px 12px; color:#4A4238; white-space:pre-wrap; margin-bottom:8px;'>"
                                        f"{_rev}</div>", unsafe_allow_html=True)
                                _priv = (r.get("VIP私訊回覆", "") or "").strip()
                                if _priv:
                                    st.markdown("**💌 要私訊的內容（點右上角圖示即可複製）：**")
                                    st.code(_priv, language="text")
                                else:
                                    st.caption("（這筆沒有私訊內容）")

                                # 🖼️ 當初的評價截圖：存在「評價截圖素材」分頁，按了才去抓（圖很大，不預先載）
                                _shk = f"dm_shot_{r['_row']}"
                                if not st.session_state.get(_shk):
                                    if st.button("🖼️ 看當初的評價截圖", key=f"dm_shotbtn_{r['_row']}",
                                                 use_container_width=True):
                                        st.session_state[_shk] = True
                                        st.rerun()
                                else:
                                    _cache = st.session_state.setdefault("_shotimgs", {})
                                    if r["_row"] not in _cache:
                                        with st.spinner("讀取截圖中…"):
                                            _sws, _srow = review_shot_find(doc, _acc, _t)
                                            _cache[r["_row"]] = (review_shot_load(_sws, _srow)
                                                                 if _sws is not None else None)
                                    _img = _cache.get(r["_row"])
                                    if _img is not None:
                                        st.image(_img, width=300)
                                    else:
                                        st.caption("找不到這筆的原始截圖（當時可能沒勾「保存截圖」，"
                                                   "或截圖存在別的帳號名下）。")
                                    if st.button("收合截圖", key=f"dm_shothide_{r['_row']}",
                                                 use_container_width=True):
                                        st.session_state.pop(_shk, None)
                                        st.rerun()

    # ==========================================
    # 📚 客服問題分類庫（獨立導航頁）：搜尋 → 複製範本；新增／匯入／管理分類都收在下方折疊區
    # ==========================================
    elif menu == "問題分類庫":
        st.markdown("""
        <div class="main-title-box">
            <div class="main-title-text">✦ 客服問題分類庫 ✦</div>
        </div>
        """, unsafe_allow_html=True)

        try:
            qa_ws, qa_rows = qa_load(doc)
        except Exception as e:
            qa_ws, qa_rows = None, []
            st.error(f"讀取問題庫失敗：{e}")

        if qa_ws is not None:
            qa_cats, qa_cfg_ws = qa_cats_load(doc)
            _used = sorted({(r["分類"] or "").strip() for r in qa_rows if (r["分類"] or "").strip()})
            all_cats = qa_cats + [c for c in _used if c not in qa_cats]   # 舊資料用過的分類也要留著

            def _cat_picker(pfx, current=""):
                """分類選單：清單裡挑，或選最後一項自己打新的。回傳最後決定的分類字串。"""
                opts = all_cats + ["➕ 自己打新分類"]
                idx = opts.index(current) if current in opts else (len(opts) - 1 if current else 0)
                sel = st.selectbox("分類", opts, index=idx, key=f"{pfx}_sel")
                if sel.startswith("➕"):
                    return st.text_input("新分類名稱", value=("" if current in all_cats else current),
                                         key=f"{pfx}_newcat", placeholder="例如 包裝破損").strip()
                return sel

            # 🔍 搜尋（這頁的主角，放最上面）
            c_s, c_f = st.columns([2, 1])
            kw = c_s.text_input("搜尋", key="qa_kw", label_visibility="collapsed",
                                placeholder="🔍 打客人問的關鍵字，例如 退貨、物流、幾天到")
            pick = c_f.selectbox("分類篩選", ["全部"] + all_cats, key="qa_filter",
                                 label_visibility="collapsed")

            def _qa_match(r):
                if pick != "全部" and (r["分類"] or "").strip() != pick:
                    return False
                kws = (kw or "").lower().split()
                if kws:
                    blob = " ".join([r["分類"], r["問題標題"], r["客戶問題範例"],
                                     r["建議回覆範本"], r["關鍵字"]]).lower()
                    return all(t in blob for t in kws)
                return True

            shown = [r for r in qa_rows if _qa_match(r)]
            shown.sort(key=lambda r: r.get("更新時間", ""), reverse=True)
            st.caption(f"共 {len(qa_rows)} 筆，符合條件 {len(shown)} 筆")

            if not qa_rows:
                st.info("問題庫還是空的。往下用「➕ 新增問題範本」建一筆，"
                        "或用「📥 從蝦皮客服對話批次匯入」讓 AI 幫你整理。")
            elif not shown:
                st.info("找不到符合的範本，換個關鍵字或把分類切回「全部」。")

            for r in shown:
                with st.container(border=True):
                    if st.session_state.get("qa_edit") == r["_row"]:
                        # ✏️ 編輯模式
                        _rw = r["_row"]
                        ecat = _cat_picker(f"qa_e_cat_{_rw}", r["分類"].strip())
                        etitle = st.text_input("問題標題", key=f"qa_e_title_{_rw}")
                        eq = st.text_area("客戶問題範例", key=f"qa_e_q_{_rw}", height=80)
                        # 🔑 回覆框用「代次 nonce」當 key：AI 改寫後換一把 key 重建，內容才換得掉。
                        #    （元件建立後不能再改它的 session_state，所以不能直接寫回同一把 key。）
                        _ern = int(st.session_state.get(f"qa_e_rnonce_{_rw}", 0))
                        # 📏 回覆框高度依內容自動加長：整段看得到才好編輯（不用一直捲）
                        _rval = st.session_state.get(f"qa_e_rval_{_rw}", r["建議回覆範本"])
                        ereply = st.text_area("建議回覆範本（回覆客戶的訊息）",
                                              height=qa_reply_height(_rval),
                                              key=f"qa_e_reply_{_rw}_{_ern}", value=_rval)
                        # ✍️ 想加強什麼 → AI 套進整則回覆
                        _eins = st.text_input("✍️ 想加強／調整什麼？（AI 會融進整則回覆）",
                                              key=f"qa_e_ins_{_rw}",
                                              placeholder="例如：多強調七天鑑賞期、語氣再親切一點、補一句提醒保留外箱")
                        e_img = st.file_uploader("📷 客戶問題截圖（選填，AI 會一起讀）",
                                                 type=["png", "jpg", "jpeg"], key=f"qa_e_img_{_rw}")
                        ekw = st.text_input("關鍵字（空格分隔）", key=f"qa_e_kw_{_rw}")

                        a1, a2 = st.columns(2)
                        _do_refine = a1.button("🤖 依指示改寫回覆", key=f"qa_e_ref_{_rw}",
                                               use_container_width=True)
                        _do_regen = a2.button("🔄 整則重新生成", key=f"qa_e_ai_{_rw}",
                                              use_container_width=True,
                                              help="不看原本的回覆，依「客戶問題範例」重寫一則新的")
                        b1, b2 = st.columns(2)
                        _do_save = b1.button("💾 儲存", key=f"qa_e_save_{_rw}", type="primary",
                                             use_container_width=True)
                        _do_cancel = b2.button("✖ 取消", key=f"qa_e_cancel_{_rw}", use_container_width=True)

                        def _put_reply(_txt, _row=_rw, _n=_ern):
                            """把 AI 結果放進回覆框（換 key 重建），並清掉加強指示。"""
                            st.session_state[f"qa_e_rval_{_row}"] = _txt
                            st.session_state[f"qa_e_rnonce_{_row}"] = _n + 1

                        if _do_refine or _do_regen:
                            _eimg = _pil_from_upload(e_img)
                            if not api_key:
                                st.warning("尚未設定 API 金鑰。")
                            elif _do_refine and not _eins.strip():
                                st.warning("請先在「想加強／調整什麼」寫一句話，AI 才知道要往哪改。")
                            elif _do_regen and not (eq.strip() or _eimg is not None):
                                st.warning("請先填客戶問題範例或上傳截圖。")
                            else:
                                with st.spinner("AI 生成中…"):
                                    if _do_refine:
                                        out, usage = qa_ai_refine(
                                            api_key, ereply, _eins, category=ecat, question=eq,
                                            model=st.session_state.get("qa_model", GEMINI_DEFAULT),
                                            image=_eimg)
                                    else:
                                        out, usage = qa_ai_suggest(
                                            api_key, ecat, eq,
                                            model=st.session_state.get("qa_model", GEMINI_DEFAULT),
                                            image=_eimg)
                                ai_track_cost(usage)
                                if out:
                                    _put_reply(out)
                                    st.rerun()
                                else:
                                    st.error("AI 生成失敗，請稍後再試。")
                        if _do_save:
                            rec = {"ID": r["ID"] or datetime.now().strftime("%Y%m%d%H%M%S%f"),
                                   "分類": ecat.strip(), "問題標題": etitle.strip(),
                                   "客戶問題範例": eq.strip(), "建議回覆範本": ereply.strip(),
                                   "關鍵字": ekw.strip(),
                                   "更新時間": datetime.now().strftime("%Y-%m-%d %H:%M")}
                            try:
                                qa_update(qa_ws, _rw, rec)
                                for _k in ("qa_edit",):
                                    st.session_state.pop(_k, None)
                                st.session_state.pop(f"qa_e_rval_{_rw}", None)
                                st.session_state.pop(f"qa_e_rnonce_{_rw}", None)
                                st.success("已更新 ✅")
                                st.rerun()
                            except Exception as e:
                                st.error(f"更新失敗：{e}")
                        if _do_cancel:
                            st.session_state.pop("qa_edit", None)
                            st.session_state.pop(f"qa_e_rval_{_rw}", None)
                            st.session_state.pop(f"qa_e_rnonce_{_rw}", None)
                            st.rerun()
                    else:
                        # 👁️ 檢視模式：標題＋分類標籤＋可複製的完整範本
                        tag = f"`{r['分類']}`　" if (r["分類"] or "").strip() else ""
                        st.markdown(f"#### {tag}{r['問題標題'] or '（未命名問題）'}")
                        # 📂 內容改成「下拉展開才看」：清單保持清爽，要用哪一則再點開複製
                        _preview = " ".join(str(r["建議回覆範本"] or "").split())[:36]
                        with st.expander(f"👀 展開看完整內容／複製　{('｜' + _preview + '…') if _preview else ''}",
                                         expanded=False):
                            if (r["客戶問題範例"] or "").strip():
                                st.markdown(f"<div style='color:#798571; margin-bottom:6px;'>💬 "
                                            f"{r['客戶問題範例']}</div>", unsafe_allow_html=True)
                            if (r["建議回覆範本"] or "").strip():
                                st.code(r["建議回覆範本"], language="text")
                            else:
                                st.caption("這筆還沒有建議回覆範本。")
                        _c1, _c2, _c3 = st.columns([2, 1, 1])
                        _meta = " · ".join([x for x in [(r["關鍵字"] or "").strip(),
                                                        (r["更新時間"] or "").strip()] if x])
                        if _meta:
                            _c1.caption(_meta)
                        if _c2.button("✏️ 編輯", key=f"qa_edit_btn_{r['_row']}", use_container_width=True):
                            st.session_state.qa_edit = r["_row"]
                            st.session_state[f"qa_e_title_{r['_row']}"] = r["問題標題"]
                            st.session_state[f"qa_e_q_{r['_row']}"] = r["客戶問題範例"]
                            st.session_state[f"qa_e_kw_{r['_row']}"] = r["關鍵字"]
                            # 回覆框走 rval＋nonce（見編輯區說明）：每次進編輯都從雲端內容乾淨起步
                            st.session_state[f"qa_e_rval_{r['_row']}"] = r["建議回覆範本"]
                            st.session_state[f"qa_e_rnonce_{r['_row']}"] = \
                                int(st.session_state.get(f"qa_e_rnonce_{r['_row']}", 0)) + 1
                            st.rerun()
                        if st.session_state.get("qa_del") == r["_row"]:
                            st.warning("確定要刪除這筆範本嗎？刪了無法復原。")
                            d1, d2 = st.columns(2)
                            if d1.button("🗑️ 確定刪除", key=f"qa_del_yes_{r['_row']}",
                                         use_container_width=True):
                                try:
                                    qa_delete(qa_ws, r["_row"])
                                    st.session_state.pop("qa_del", None)
                                    st.success("已刪除 ✅")
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"刪除失敗：{e}")
                            if d2.button("取消", key=f"qa_del_no_{r['_row']}", use_container_width=True):
                                st.session_state.pop("qa_del", None)
                                st.rerun()
                        else:
                            if _c3.button("🗑️ 刪除", key=f"qa_del_btn_{r['_row']}", use_container_width=True):
                                st.session_state.qa_del = r["_row"]
                                st.rerun()

            st.divider()

            # ➕ 新增問題範本
            with st.expander("➕ 新增問題範本", expanded=not qa_rows):
                _ncat = _cat_picker("qa_new_cat")
                st.text_input("問題標題", key="qa_new_title", placeholder="例如 包裹遲遲未到")
                st.text_area("客戶問題範例", key="qa_new_q", height=80,
                             placeholder="客人實際會怎麼問（也是 AI 生成的依據；可只上傳截圖）")
                qa_new_img = st.file_uploader("📷 客戶問題截圖（選填，AI 會一起讀圖內容）",
                                              type=["png", "jpg", "jpeg"], key="qa_new_img")
                # 回覆框同樣走 rval＋nonce，AI 生成／改寫後才換得掉內容
                _nrn = int(st.session_state.get("qa_new_rnonce", 0))
                _new_reply = st.text_area("建議回覆範本（回覆客戶的訊息，可直接編輯）",
                                          height=qa_reply_height(
                                              st.session_state.get("qa_new_rval", "")),
                                          key=f"qa_new_reply_{_nrn}",
                                          value=st.session_state.get("qa_new_rval", ""),
                                          placeholder="可手動輸入，或用下面的 AI 生成後再微調")
                _nins = st.text_input("✍️ 想加強／調整什麼？（AI 會融進整則回覆）", key="qa_new_ins",
                                      placeholder="例如：多強調七天鑑賞期、語氣再親切一點、補一句提醒保留外箱")
                n1, n2 = st.columns(2)
                _n_refine = n1.button("🤖 依指示改寫回覆", key="qa_new_ref", use_container_width=True)
                _n_gen = n2.button("🔄 整則重新生成", key="qa_new_ai", use_container_width=True,
                                   help="不看現有內容，依「客戶問題範例」寫一則新的")
                if _n_refine or _n_gen:
                    _img = _pil_from_upload(qa_new_img)
                    _q_now = st.session_state.get("qa_new_q", "")
                    if not api_key:
                        st.warning("尚未設定 API 金鑰，無法使用 AI。")
                    elif _n_refine and not _nins.strip():
                        st.warning("請先在「想加強／調整什麼」寫一句話，AI 才知道要往哪改。")
                    elif _n_gen and not (_q_now.strip() or _img is not None):
                        st.warning("請先填「客戶問題範例」或上傳截圖，AI 才知道要回什麼。")
                    else:
                        with st.spinner("AI 生成中…"):
                            if _n_refine:
                                out, usage = qa_ai_refine(
                                    api_key, _new_reply, _nins, category=_ncat, question=_q_now,
                                    model=st.session_state.get("qa_model", GEMINI_DEFAULT), image=_img)
                            else:
                                out, usage = qa_ai_suggest(
                                    api_key, _ncat, _q_now,
                                    model=st.session_state.get("qa_model", GEMINI_DEFAULT), image=_img)
                        ai_track_cost(usage)
                        if out:
                            st.session_state["qa_new_rval"] = out
                            st.session_state["qa_new_rnonce"] = _nrn + 1
                            st.rerun()
                        else:
                            st.error("AI 生成失敗，請稍後再試。")
                st.text_input("關鍵字（用空格分隔，方便日後搜尋）", key="qa_new_kw",
                              placeholder="例如 退貨 七天 鑑賞期")
                if st.button("💾 儲存到問題庫", key="qa_new_save", type="primary", use_container_width=True):
                    _title = st.session_state.get("qa_new_title", "").strip()
                    _q = st.session_state.get("qa_new_q", "").strip()
                    if not (_title or _q):
                        st.warning("至少要填「問題標題」或「客戶問題範例」其中一項。")
                    else:
                        rec = {"ID": datetime.now().strftime("%Y%m%d%H%M%S%f"),
                               "分類": _ncat.strip(), "問題標題": _title, "客戶問題範例": _q,
                               "建議回覆範本": (_new_reply or "").strip(),
                               "關鍵字": st.session_state.get("qa_new_kw", "").strip(),
                               "更新時間": datetime.now().strftime("%Y-%m-%d %H:%M")}
                        try:
                            qa_add(qa_ws, rec)
                            for k in ["qa_new_title", "qa_new_q", "qa_new_kw", "qa_new_ins", "qa_new_rval"]:
                                st.session_state.pop(k, None)
                            st.session_state["qa_new_rnonce"] = _nrn + 1   # 清空回覆框
                            st.success("已新增到問題庫 ✅")
                            st.rerun()
                        except Exception as e:
                            st.error(f"儲存失敗：{e}")

            # 📥 從蝦皮客服對話批次匯入
            with st.expander("📥 從蝦皮客服對話批次匯入（AI 自動分類）", expanded=False):
                st.caption("把蝦皮聊聊的對話整段貼進來（或上傳對話截圖），AI 會拆成「一個問題一筆」，"
                           "回覆以你當時實際回過的內容為準做潤飾，個資會自動拿掉。")
                imp_text = st.text_area("貼上蝦皮客服對話", key="qa_imp_text", height=170,
                                        placeholder="客人：請問這個發貨要幾天？\n我：您好～目前備貨約 1-2 個工作天…")
                imp_imgs = st.file_uploader("📷 或上傳對話截圖（可多張）", type=["png", "jpg", "jpeg"],
                                            accept_multiple_files=True, key="qa_imp_imgs")
                imp_hint = st.text_input("補充說明（選填，例如：這是同一位客人的退貨案）", key="qa_imp_hint")
                if st.button("🤖 分析並整理成問題範本", key="qa_imp_run", use_container_width=True):
                    _imgs = [i for i in [_pil_from_upload(f) for f in (imp_imgs or [])] if i is not None]
                    if not api_key:
                        st.warning("尚未設定 API 金鑰，無法使用 AI。")
                    elif not ((imp_text or "").strip() or _imgs):
                        st.warning("請先貼上對話內容或上傳截圖。")
                    else:
                        with st.spinner("AI 整理中…（對話越長越久，請稍候）"):
                            items, usage = qa_ai_from_chat(
                                api_key, imp_text, model=st.session_state.get("qa_model", GEMINI_DEFAULT),
                                images=_imgs, hint=imp_hint or "")
                        ai_track_cost(usage)
                        if items:
                            st.session_state["qa_imp_items"] = items
                            for k in [k for k in list(st.session_state.keys()) if str(k).startswith("qa_ip_")]:
                                st.session_state.pop(k, None)
                            st.rerun()
                        else:
                            st.error("AI 沒有整理出結果（可能對話太短或格式看不懂），可以再貼一次或改用截圖試試。")

                _imp_items = st.session_state.get("qa_imp_items") or []
                if _imp_items:
                    _exist_titles = {(r["問題標題"] or "").strip() for r in qa_rows}
                    st.success(f"AI 整理出 {len(_imp_items)} 筆，確認內容後再存入（可直接修改）：")
                    for _i, _it in enumerate(_imp_items):
                        _dup = _it["問題標題"].strip() != "" and _it["問題標題"].strip() in _exist_titles
                        st.session_state.setdefault(f"qa_ip_use_{_i}", not _dup)
                        for _f, _k in (("分類", "cat"), ("問題標題", "title"), ("客戶問題範例", "q"),
                                       ("建議回覆範本", "reply"), ("關鍵字", "kw")):
                            st.session_state.setdefault(f"qa_ip_{_k}_{_i}", _it[_f])
                        with st.container(border=True):
                            st.checkbox(f"{'⚠️ 問題庫已有同名範本：' if _dup else ''}"
                                        f"{_it['問題標題'] or '（未命名）'}", key=f"qa_ip_use_{_i}")
                            p1, p2 = st.columns([1, 2])
                            p1.text_input("分類", key=f"qa_ip_cat_{_i}")
                            p2.text_input("問題標題", key=f"qa_ip_title_{_i}")
                            st.text_area("客戶問題範例", key=f"qa_ip_q_{_i}", height=70)
                            st.text_area("建議回覆範本", key=f"qa_ip_reply_{_i}", height=260)
                            st.text_input("關鍵字", key=f"qa_ip_kw_{_i}")
                    v1, v2 = st.columns(2)
                    if v1.button("💾 把勾選的存進問題庫", key="qa_imp_save", type="primary",
                                 use_container_width=True):
                        _new = []
                        for _i in range(len(_imp_items)):
                            if not st.session_state.get(f"qa_ip_use_{_i}"):
                                continue
                            _ti = st.session_state.get(f"qa_ip_title_{_i}", "").strip()
                            _qi = st.session_state.get(f"qa_ip_q_{_i}", "").strip()
                            if not (_ti or _qi):
                                continue
                            _new.append([datetime.now().strftime("%Y%m%d%H%M%S%f") + str(_i),
                                         st.session_state.get(f"qa_ip_cat_{_i}", "").strip(), _ti, _qi,
                                         st.session_state.get(f"qa_ip_reply_{_i}", "").strip(),
                                         st.session_state.get(f"qa_ip_kw_{_i}", "").strip(),
                                         datetime.now().strftime("%Y-%m-%d %H:%M")])
                        if not _new:
                            st.warning("沒有勾選任何一筆（或勾選的都沒填標題／問題）。")
                        else:
                            try:
                                qa_ws.append_rows(_new, value_input_option="RAW")
                                st.session_state.pop("qa_imp_items", None)
                                st.session_state.pop("qa_imp_text", None)
                                for _k in [k for k in list(st.session_state.keys())
                                           if str(k).startswith("qa_ip_")]:
                                    st.session_state.pop(_k, None)
                                st.success(f"已存入 {len(_new)} 筆 ✅")
                                st.rerun()
                            except Exception as e:
                                st.error(f"存入失敗：{e}")
                    if v2.button("✖ 丟掉這批結果", key="qa_imp_drop", use_container_width=True):
                        st.session_state.pop("qa_imp_items", None)
                        for _k in [k for k in list(st.session_state.keys()) if str(k).startswith("qa_ip_")]:
                            st.session_state.pop(_k, None)
                        st.rerun()

            # ⚙️ 管理分類：一行一個，想加想改想刪都在這裡
            with st.expander("⚙️ 管理分類（想加、想改名、想刪都在這）", expanded=False):
                st.caption("一行一個分類，改完按儲存。已經用到這個分類的舊資料不會被改動，"
                           "所以改名後舊資料仍會保留原本的分類名（會自動出現在選單裡）。")
                # 🔑 這個框用「代次 nonce」當 key：存檔後換一把 key 重建，框裡才會顯示存好的內容。
                #    （不可以在元件建立後才去改它的 session_state，Streamlit 會直接報錯。）
                _cn = int(st.session_state.get("qa_cats_nonce", 0))
                _cats_txt = st.text_area("分類清單", value="\n".join(qa_cats), height=210,
                                         key=f"qa_cats_txt_{_cn}")
                g1, g2 = st.columns(2)
                if g1.button("💾 儲存分類", key="qa_cats_save", type="primary", use_container_width=True):
                    if qa_cfg_ws is None:
                        st.error("連不上「系統設定」分頁，無法儲存。")
                    else:
                        try:
                            _saved = qa_cats_save(qa_cfg_ws, (_cats_txt or "").split("\n"))
                            st.session_state["qa_cats_nonce"] = _cn + 1
                            st.success(f"已儲存 {len(_saved)} 個分類 ✅")
                            st.rerun()
                        except Exception as e:
                            st.error(f"儲存失敗：{e}")
                if g2.button("↩ 回到預設分類", key="qa_cats_reset", use_container_width=True):
                    if qa_cfg_ws is None:
                        st.error("連不上「系統設定」分頁，無法儲存。")
                    else:
                        try:
                            qa_cats_save(qa_cfg_ws, QA_DEFAULT_CATS)
                            st.session_state["qa_cats_nonce"] = _cn + 1
                            st.success("已還原成預設分類 ✅")
                            st.rerun()
                        except Exception as e:
                            st.error(f"還原失敗：{e}")

            # 🤖 AI 設定（不常動，收在最後，不要一開頁就一大排）
            with st.expander("🤖 AI 模型與花費（預設免費款）", expanded=False):
                _mkeys = list(GEMINI_PRICES.keys())
                st.selectbox("AI 模型", _mkeys, index=_mkeys.index(GEMINI_DEFAULT),
                             format_func=lambda k: GEMINI_PRICES[k]["label"], key="qa_model",
                             help="預設為免費額度的推薦款；標「付費」的模型才會真的扣費。")
                ai_render_cost(st.session_state.get("qa_model", GEMINI_DEFAULT))

    # ==========================================
    # ✨ 動態文字壓印版：折價券管理
    # ==========================================
    elif menu == "折價券管理":
        st.markdown("""
        <div class="main-title-box">
            <div class="main-title-text">✦ 動態日期折價券管理 ✦</div>
        </div>
        """, unsafe_allow_html=True)
        
        ws_cfg = get_or_create_ws(doc, "系統設定")
        
        if "cfg_data" not in st.session_state or st.session_state.get("refresh_cfg", True):
            st.session_state.cfg_data = ws_cfg.get_all_values()
            st.session_state.refresh_cfg = False
        cfg_data = st.session_state.cfg_data
        
        if "active_slots" not in st.session_state:
            order_row = next((r for r in cfg_data if len(r) > 1 and r[0] == 'coupon_order'), None)
            if order_row:
                raw_slots = [int(x) for x in order_row[1].split(",") if x]
                st.session_state.active_slots = list(dict.fromkeys(raw_slots))
            else:
                existing_ids = [int(r[0].split('_')[1]) for r in cfg_data if len(r) > 0 and str(r[0]).startswith('coupon_')]
                st.session_state.active_slots = sorted(list(set(existing_ids))) if existing_ids else [1, 2]
            
        c_add, c_space = st.columns([2, 5])
        with c_add:
            if st.button("➕ 新增一個全新版位", type="primary", use_container_width=True):
                next_id = 1
                while next_id in st.session_state.active_slots:
                    next_id += 1
                st.session_state.active_slots.append(next_id)
                trigger_order_save(sheet_url, st.session_state.active_slots)
                st.rerun()
        
        now = datetime.now()
        last_day = calendar.monthrange(now.year, now.month)[1]
        default_coupon_txt = f"{now.month}/{last_day}"
        
        st.markdown("<div style='margin-bottom: 5px;'></div>", unsafe_allow_html=True)
        
        display_idx = 0
        for slot_id in st.session_state.active_slots:
            idx = st.session_state.active_slots.index(slot_id)
            is_first = (idx == 0)
            is_last = (idx == len(st.session_state.active_slots) - 1)
            display_num = display_idx + 1 
            
            if display_idx % 2 == 0:
                col1, col2 = st.columns(2, gap="medium")
                current_col = col1
            else:
                current_col = col2
            display_idx += 1
                
            with current_col:
                # 埋入隱形錨點：讓 CSS 在手機上把左右兩欄改成上下單欄
                st.markdown('<span class="coupon-grid-anchor" style="display:none;"></span>', unsafe_allow_html=True)

                # 🔒 保護鎖：鎖住的版位刪不掉（避免誤按 ✕ 把常用的那張券刪掉）
                _lock_row = next((r for r in cfg_data if len(r) > 1 and r[0] == f'couponlock_{slot_id}'), None)
                _locked = bool(_lock_row and str(_lock_row[1]).strip() == "1")

                # ✨ 標題＋上移／下移／刪除／保護鎖：同一排（標題在左，四顆方塊鍵靠右）
                c_title, c_up, c_down, c_del, c_lock = st.columns(5)
                with c_title:
                    # 埋入隱形錨點供 CSS 辨識（第一欄彈性、後三欄固定 40px 方塊鍵）
                    st.markdown('<span class="inline-row-btn" style="display:none;"></span>', unsafe_allow_html=True)
                    st.markdown(f"<p class='sub-title-text' style='margin:0; line-height:40px;'>折價券版位 {display_num}</p>", unsafe_allow_html=True)
                with c_up:
                    if st.button("↑", key=f"up_btn_{slot_id}", disabled=is_first):
                        st.session_state.active_slots[idx], st.session_state.active_slots[idx-1] = st.session_state.active_slots[idx-1], st.session_state.active_slots[idx]
                        trigger_order_save(sheet_url, st.session_state.active_slots)
                        st.rerun()
                with c_down:
                    if st.button("↓", key=f"dn_btn_{slot_id}", disabled=is_last):
                        st.session_state.active_slots[idx], st.session_state.active_slots[idx+1] = st.session_state.active_slots[idx+1], st.session_state.active_slots[idx]
                        trigger_order_save(sheet_url, st.session_state.active_slots)
                        st.rerun()
                with c_del:
                    if st.button("✕", key=f"del_btn_{slot_id}", disabled=_locked,
                                 help="刪掉這個版位（鎖住時按不動）"):
                        rows_to_del = [i + 1 for i, r in enumerate(cfg_data) if r and r[0] == f"coupon_{slot_id}"]
                        if rows_to_del:
                            with st.spinner("刪除中..."):
                                for row_index in sorted(rows_to_del, reverse=True):
                                    ws_cfg.delete_rows(row_index)
                                st.session_state.refresh_cfg = True
                        st.session_state.active_slots.remove(slot_id)
                        trigger_order_save(sheet_url, st.session_state.active_slots)
                        st.rerun()
                with c_lock:
                    if st.button("🔒" if _locked else "🔓", key=f"lock_btn_{slot_id}",
                                 help=("已上鎖：這個版位刪不掉。按一下解鎖。" if _locked
                                       else "按一下上鎖：鎖住後就不會被誤刪。")):
                        _save_kv(ws_cfg, f"couponlock_{slot_id}", "" if _locked else "1")
                        st.session_state.refresh_cfg = True
                        st.rerun()

                # 🧼 關鍵修正：分開「乾淨底圖」與「成品圖」，壓字一律壓在乾淨底圖上，不會疊字。
                #    base_img = 乾淨底圖(couponbase_)；display_img = 已存成品(coupon_，含字，給預覽/下載)
                base_img = None
                display_img = None
                existing_row = next((r for r in cfg_data if len(r) > 0 and r[0] == f'coupon_{slot_id}'), None)
                clean_row = next((r for r in cfg_data if len(r) > 0 and r[0] == f'couponbase_{slot_id}'), None)
                cache = st.session_state.setdefault("_decoded_imgs", {})

                def _decode_row(row, ck):
                    if not (row and len(row) > 1):
                        return None
                    sig = (len(row), row[1][:32])
                    ent = cache.get(ck)
                    if ent and ent[0] == sig:
                        return ent[1]
                    img = base64_chunks_to_img([c for c in row[1:] if c])
                    cache[ck] = (sig, img)
                    return img

                try:
                    display_img = _decode_row(existing_row, f"disp_{slot_id}")
                    base_img = _decode_row(clean_row, f"base_{slot_id}")
                    if base_img is None:
                        base_img = display_img      # 舊資料沒有乾淨底圖 → 暫用成品(可能已含字)
                    if display_img is None:
                        display_img = base_img
                    if display_img is not None:
                        # 📱 預覽改用「完整原圖」當 src、只用 CSS 顯示縮小：
                        #    這樣手機長按「加入照片」存進相簿的是完整畫質（1080），不再是 300px 糊圖
                        _pv = BytesIO()
                        display_img.save(_pv, format="PNG")
                        _pv_b64 = base64.b64encode(_pv.getvalue()).decode()
                        st.markdown(
                            f'<img src="data:image/png;base64,{_pv_b64}" '
                            f'style="width:300px; max-width:100%; height:auto; border-radius:10px;" '
                            f'alt="折價券{display_num}">',
                            unsafe_allow_html=True)
                    else:
                        st.info("此版位目前為空，請先上傳底圖。")
                except Exception:
                    st.warning("圖片載入異常，請重新上傳")

                # ✨ 下載鈕＋提示改成上下兩行：並排在手機上會擠到重疊、右邊字還會被切掉
                if base_img:
                    buf = BytesIO()
                    # ✨ 畫質升級：無損 PNG 下載（下載已存成品＝含字那張）
                    (display_img or base_img).save(buf, format="PNG")
                    st.download_button(label="💻 下載到電腦", data=buf.getvalue(),
                                       file_name=f"BearJoy_Coupon_{display_num}.png", mime="image/png",
                                       key=f"dl_btn_{slot_id}", use_container_width=True)
                    st.caption("💡 手機直接長按上面那張圖 → 存到相簿（原畫質）")

                    # 📅 快速換日期：打上日期，就用「上次記住的位置／大小／顏色」壓在乾淨底圖上直接出券。
                    #    刻意「不寫回雲端」→ 雲端那張永遠是空白底圖，不會被上個月的日期蓋掉，隨時可再換。
                    _sv_row = next((r for r in cfg_data if len(r) > 1 and r[0] == f'coupset_{slot_id}'), None)
                    _sv = _parse_coupset(_sv_row)
                    with st.expander("📅 快速換日期（打上日期直接出券）", expanded=bool(_sv)):
                        if clean_row is None:
                            st.caption("⚠️ 這個版位還沒存過『乾淨底圖』，壓出來可能會疊到舊的字。"
                                       "建議重新上傳一次沒有字的底圖並按「✅ 直接儲存原圖（不加字）」。")
                        if not _sv:
                            st.caption("還沒記住日期要壓在哪裡。先用下面「✏️ 想加日期/文字」喬好位置、按「✅ 確認儲存」一次，"
                                       "之後每次就能在這裡直接換日期。")
                        _qd = st.text_input("日期", value="", key=f"qd_{slot_id}",
                                            placeholder=f"例如 {default_coupon_txt}",
                                            help="打完馬上出券（用上次記住的位置、大小、顏色）。雲端底圖不會被改動。")
                        if (_qd or "").strip() and _sv:
                            _qimg, _, _ = _stamp_coupon(
                                base_img, _qd.strip(), _sv.get("color", "#FFFFFF"),
                                int(_sv.get("size", 50)),
                                int(_sv.get("x", base_img.width // 2)),
                                int(_sv.get("y", int(base_img.height * 0.7))),
                                int(_sv.get("rot", 0)))
                            _qbuf = BytesIO()
                            _qimg.save(_qbuf, format="PNG")
                            st.markdown(
                                f'<img src="data:image/png;base64,'
                                f'{base64.b64encode(_qbuf.getvalue()).decode()}" '
                                f'style="width:300px; max-width:100%; height:auto; border-radius:10px;" '
                                f'alt="折價券{display_num}">', unsafe_allow_html=True)
                            st.download_button("💻 下載這張（含日期）", data=_qbuf.getvalue(),
                                               file_name=f"BearJoy_Coupon_{display_num}_"
                                                         f"{_safe_filename(_qd.strip())}.png",
                                               mime="image/png", key=f"qdl_{slot_id}",
                                               use_container_width=True)
                            st.caption("💡 手機長按上圖 → 存到相簿。雲端那張仍是空白底圖，下次可以再換別的日期。")

                new_file = st.file_uploader(f"更換版位 {display_num} 圖片", type=["png", "jpg", "jpeg"], key=f"up_file_{slot_id}", label_visibility="collapsed")
                
                if new_file:
                    base_img = Image.open(new_file)
                    # 直接上傳、不需編輯：一鍵存原圖
                    if st.button("✅ 直接儲存原圖（不加字）", type="primary", use_container_width=True, key=f"save_raw_{slot_id}"):
                        with st.spinner("儲存中..."):
                            chunks = img_to_base64_chunks(base_img.convert("RGB"))
                            # 新底圖：成品(coupon_)與乾淨底圖(couponbase_)都存這張、並清掉舊位置記憶(coupset_)
                            _keys = (f"coupon_{slot_id}", f"couponbase_{slot_id}", f"coupset_{slot_id}")
                            for ri in sorted([i + 1 for i, r in enumerate(cfg_data) if r and r[0] in _keys], reverse=True):
                                ws_cfg.delete_rows(ri)
                            ws_cfg.append_row([f"coupon_{slot_id}"] + chunks)
                            ws_cfg.append_row([f"couponbase_{slot_id}"] + chunks)
                            st.session_state.pop("_decoded_imgs", None)
                            st.session_state.refresh_cfg = True
                            st.success("已儲存！")
                            st.rerun()

                if base_img:
                    with st.expander("✏️ 想加日期/文字再點開（不需要可略過）", expanded=False):
                        enable_text = st.checkbox("✒️ 啟動文字壓印", value=True, key=f"en_txt_{slot_id}")
                        final_img_to_save = base_img 
                        
                        if enable_text:
                            # 📅 功能3：讀取此版位上次存的壓印位置，換日期就不用再喬位置
                            sv_row = next((r for r in cfg_data if len(r) > 1 and r[0] == f'coupset_{slot_id}'), None)
                            sv = _parse_coupset(sv_row)
                            def_size = max(10, min(sv.get("size", 50), 200))
                            def_rot = max(-180, min(sv.get("rot", 0), 180))
                            def_color = sv.get("color", "#FFFFFF")
                            def_x = max(0, min(sv.get("x", base_img.width // 2), base_img.width))
                            def_y = max(0, min(sv.get("y", int(base_img.height * 0.7)), base_img.height))
                            st.caption("在下方「文字」打上日期，文字會出現在畫布，再用拖曳調整位置；不打字就不會壓任何文字。")

                            # ✨ 錨點魔法 2：文字方框與顏色並排
                            c_txt, c_col = st.columns(2)
                            with c_txt:
                                # 埋入隱形錨點供 CSS 辨識
                                st.markdown('<span class="inline-row-txt" style="display:none;"></span>', unsafe_allow_html=True)
                                # label 要短：跟右邊顏色欄並排時，長 label 會被壓到重疊
                                text_input = st.text_input("✍️ 文字", value="",
                                                           placeholder=f"例如 {default_coupon_txt}",
                                                           key=f"txt_{slot_id}",
                                                           help="預設空白＝不壓任何文字；要壓日期就自己打，例如 6/30")
                            with c_col:
                                text_color = st.color_picker("🎨 顏色", def_color, key=f"col_{slot_id}")

                            # 操作方式：🎨 畫布直接編輯（Canva 式）／三模式拖曳／拉桿
                            x_pos = max(0, min(int(st.session_state.get(f"px_{slot_id}", def_x)), base_img.width))
                            y_pos = max(0, min(int(st.session_state.get(f"py_{slot_id}", def_y)), base_img.height))
                            use_canvas = HAS_CANVAS  # Canva 式畫布（手機可用就用這個；按「編輯文字」才開）
                            if use_canvas or HAS_IMG_COORDS:
                                font_size = max(10, min(int(st.session_state.get(f"csz_{slot_id}", def_size)), 200))
                                rotation_angle = max(-180, min(int(st.session_state.get(f"crot_{slot_id}", def_rot)), 180))
                            else:
                                c_sz, c_rot = st.columns(2)
                                c_sz.markdown('<span class="slider-pair-anchor" style="display:none;"></span>', unsafe_allow_html=True)
                                font_size = c_sz.slider("📐 大小", 10, 200, def_size, key=f"sz_{slot_id}")
                                rotation_angle = c_rot.slider("🔄 旋轉", -180, 180, def_rot, key=f"rot_{slot_id}")
                                cx2, cy2 = st.columns(2)
                                cx2.markdown('<span class="slider-pair-anchor" style="display:none;"></span>', unsafe_allow_html=True)
                                x_pos = cx2.slider("↔️ 左右", 0, base_img.width, x_pos, key=f"x_{slot_id}")
                                y_pos = cy2.slider("↕️ 上下", 0, base_img.height, y_pos, key=f"y_{slot_id}")

                            final_img_to_save, text_w, text_h = _stamp_coupon(
                                base_img, text_input, text_color, font_size, x_pos, y_pos, rotation_angle)

                            editor_shown = False
                            if use_canvas:
                                if st.session_state.get("edit_slot") != slot_id:
                                    # 未編輯此版位：顯示靜態預覽（「編輯文字」鈕在下方與確認儲存併排）
                                    editor_shown = True
                                    st.image(final_img_to_save, width=300)
                                else:
                                    # 畫布寬度固定 px（drawable-canvas 不吃 %）；手機上 300 才不會超出容器被右邊截掉。
                                    # 螢幕夠寬（或想更好按）可自行勾「畫布加大」，圖大一點手指比較好操作。
                                    _cvbig = st.checkbox("🔍 畫布加大（若右邊被切掉就取消勾選）", key=f"cvbig_{slot_id}")
                                    disp_w = 360 if _cvbig else 300
                                    cscale = disp_w / base_img.width
                                    disp_h = max(1, int(base_img.height * cscale))
                                    bg = base_img.convert("RGB").resize((disp_w, disp_h), Image.LANCZOS)
                                    # 🔑 關鍵解法：把底圖放進 fabric 的「背景圖」欄位(backgroundImage)，
                                    # 背景圖天生就不能被選取/拖曳 → 圖永遠當底，唯一可編輯的物件只有壓印文字。
                                    # （之前用 image 物件當底，圖還是會被當成編輯對象，所以改成背景圖。）
                                    _bgbuf = BytesIO(); bg.save(_bgbuf, format="PNG")
                                    bg_uri = "data:image/png;base64," + base64.b64encode(_bgbuf.getvalue()).decode()
                                    # 🔑 只有「文字非空」才放一個可編輯文字物件；清空文字＝畫布上沒有文字（這就是刪字）。
                                    # init 的生成位置用「已提交的基準位置」(cinit*)，預設等於 def_*；
                                    # 平時保持不動（拖曳期間 init 內容不變 → 同一把 key 下畫布不會重載複製出鬼影），
                                    # 只有在下方「自動清疊字」重建畫布時，才更新成使用者拖到的位置。
                                    _init_x = max(0, min(int(st.session_state.get(f"cinitx_{slot_id}", def_x)), base_img.width))
                                    _init_y = max(0, min(int(st.session_state.get(f"cinity_{slot_id}", def_y)), base_img.height))
                                    _init_sz = max(10, min(int(st.session_state.get(f"cinitsz_{slot_id}", def_size)), 200))
                                    _init_rot = max(-180, min(int(st.session_state.get(f"cinitrot_{slot_id}", def_rot)), 180))
                                    _txt_objs = []
                                    if (text_input or "").strip():
                                        _txt_objs = [{
                                            "type": "i-text", "version": "4.4.0", "text": text_input,
                                            "left": float(_init_x * cscale), "top": float(_init_y * cscale),
                                            "originX": "center", "originY": "center",
                                            "fontSize": max(10, int(_init_sz * cscale)), "fill": text_color,
                                            "angle": float(_init_rot), "fontFamily": "sans-serif", "editable": True,
                                            # 👆 手機用：把四角縮放把手加大加粗（fabric 預設把手只有幾 px，手指幾乎按不到），
                                            #    touchCornerSize 是觸控時的實際可按範圍。
                                            "cornerSize": 22, "touchCornerSize": 48,
                                            "transparentCorners": False, "cornerStyle": "circle",
                                            "cornerColor": "#E0533D", "cornerStrokeColor": "#FFFFFF",
                                            "borderColor": "#E0533D", "borderScaleFactor": 2, "padding": 10,
                                        }]
                                    init = {
                                        "version": "4.4.0",
                                        "backgroundImage": {
                                            "type": "image", "version": "4.4.0", "src": bg_uri,
                                            "left": 0, "top": 0, "originX": "left", "originY": "top",
                                            "width": disp_w, "height": disp_h, "scaleX": 1, "scaleY": 1,
                                            "angle": 0, "opacity": 1, "crossOrigin": None, "filters": [],
                                        },
                                        "objects": _txt_objs,
                                    }
                                    canvas_ok = False
                                    # 🔑 key 隨「文字／顏色／清字代次(nonce)」變化：改字或自動清疊字時整個重建
                                    # (乾淨單一文字，不疊加)；拖曳/縮放/旋轉不改 key → 畫布不重建 → 位置保留、放開即讀回並存檔。
                                    _cnonce = int(st.session_state.get(f"cvnonce_{slot_id}", 0))
                                    _csig = f"{text_input}|{text_color}|{_cnonce}|{disp_w}"
                                    _ckey = f"cv_{slot_id}_{abs(hash(_csig))}"
                                    try:
                                        cres = st_canvas(initial_drawing=init,
                                                         # 🔑 update_streamlit=True：拖曳/縮放/旋轉一放開就把新位置回傳 Python，
                                                         # 這樣「完成編輯／確認儲存」存的才是你拖到的位置，不會跳回原位。
                                                         # display_toolbar=True：顯示工具列垃圾桶 → 可刪除選取的文字。
                                                         drawing_mode="transform", update_streamlit=True,
                                                         height=disp_h, width=disp_w, display_toolbar=True,
                                                         background_color="#FFFFFF",
                                                         key=_ckey)
                                        canvas_ok = True
                                    except Exception as _ce:
                                        cres = None
                                        st.warning(f"⚠️ 畫布無法載入（{type(_ce).__name__}），改用拖曳編輯。")
                                    if canvas_ok:
                                        editor_shown = True
                                        st.caption("✍️ 在上面「文字」框打上日期，文字就會出現在畫布。"
                                                   "拖曳＝移動、拉四角＝縮放、轉上方圓點＝旋轉、點兩下＝改字。"
                                                   "🗑️ 想刪文字：把上面「文字」框清空（或選取文字後按畫布工具列垃圾桶）。"
                                                   "喬好後按「✅ 確認儲存」（位置會即時記住，不會跳回原位）。")
                                        if cres is not None and getattr(cres, "json_data", None):
                                            objs = cres.json_data.get("objects", [])
                                            # objects[0] 現在是底圖影像，要挑出文字物件(i-text)來讀位置
                                            itexts = [ob for ob in objs if ob.get("type") == "i-text"]
                                            # 🧹 drawable-canvas 偶爾把同一段文字複製成多個物件(疊字/鬼影)。
                                            # 從中挑「使用者實際拖動」的那一個＝離 init 生成點最遠者，
                                            # 這樣移動任一個複本都會正確更新位置，不必先手動刪掉多餘的那個。
                                            _ix = float(_init_x * cscale); _iy = float(_init_y * cscale)
                                            def _moved_dist(ob):
                                                lx0, ty0 = ob.get("left"), ob.get("top")
                                                if lx0 is None or ty0 is None:
                                                    return -1.0
                                                return math.hypot(float(lx0) - _ix, float(ty0) - _iy)
                                            o = max(itexts, key=_moved_dist) if itexts else None
                                            if o:
                                                # 取 scaleX/scaleY 較大者：拉「左右邊」把手時只有其中一個會變，
                                                # 只讀 scaleX 會發生「明明拉大了字級卻沒變」。
                                                sx = max(float(o.get("scaleX", 1) or 1), float(o.get("scaleY", 1) or 1))
                                                ang = float(o.get("angle", 0) or 0)
                                                lx, ty2 = o.get("left"), o.get("top")
                                                ofs = float(o.get("fontSize", font_size * cscale) or (font_size * cscale))
                                                # ✍️ 以「畫布上實際打的文字」為準（在畫布點兩下直接編輯）
                                                _ctext = o.get("text")
                                                if isinstance(_ctext, str) and _ctext.strip():
                                                    text_input = _ctext
                                                if lx is not None and ty2 is not None:
                                                    x_pos = max(0, min(int(lx / cscale), base_img.width))
                                                    y_pos = max(0, min(int(ty2 / cscale), base_img.height))
                                                font_size = max(10, min(int(round(ofs * sx / cscale)), 200))
                                                rotation_angle = max(-180, min(int(round(((ang + 180) % 360) - 180)), 180))
                                                st.session_state[f"px_{slot_id}"] = x_pos
                                                st.session_state[f"py_{slot_id}"] = y_pos
                                                st.session_state[f"csz_{slot_id}"] = font_size
                                                st.session_state[f"crot_{slot_id}"] = rotation_angle
                                                # 🧹 自動清疊字：偵測到多個文字物件、且其中有人被真的拖走(>4px)時，
                                                # 把基準位置(cinit*)更新成「拖到的位置」並換 key 重建畫布 → 收斂成單一文字，
                                                # 不勞使用者手動刪。清完後單一物件距 init=0，不會再觸發，避免無限重跑。
                                                if len(itexts) > 1 and _moved_dist(o) > 4:
                                                    st.session_state[f"cinitx_{slot_id}"] = x_pos
                                                    st.session_state[f"cinity_{slot_id}"] = y_pos
                                                    st.session_state[f"cinitsz_{slot_id}"] = font_size
                                                    st.session_state[f"cinitrot_{slot_id}"] = rotation_angle
                                                    st.session_state[f"cvnonce_{slot_id}"] = _cnonce + 1
                                                    st.rerun()
                                            else:
                                                # 畫布上已沒有文字物件（被工具列垃圾桶刪掉）→ 存檔就不要壓任何文字。
                                                text_input = ""
                                        final_img_to_save, text_w, text_h = _stamp_coupon(
                                            base_img, text_input, text_color, font_size, x_pos, y_pos, rotation_angle)
                            if (not editor_shown) and HAS_IMG_COORDS:
                                edit_mode = st.radio("操作", ["✋ 移動", "🔍 放大縮小", "🔄 旋轉"],
                                                     horizontal=True, key=f"mode_{slot_id}", label_visibility="collapsed")
                                hint = {"✋ 移動": "在圖上按住拖曳 → 文字移到放開處",
                                        "🔍 放大縮小": "從文字往外拖＝放大、往中心拖＝縮小",
                                        "🔄 旋轉": "往哪個方向拖，文字就轉向那邊"}.get(edit_mode, "")
                                st.caption(f"目前：{edit_mode}　·　{hint}")
                                disp_w = 320
                                disp_img = final_img_to_save.resize((disp_w, max(1, int(final_img_to_save.height * disp_w / final_img_to_save.width))))
                                try:
                                    coords = st_image_coordinates(disp_img, key=f"clk_{slot_id}", click_and_drag=True)
                                except TypeError:
                                    coords = st_image_coordinates(disp_img, key=f"clk_{slot_id}")
                                if coords:
                                    rx = coords.get("x2", coords.get("x"))
                                    ry = coords.get("y2", coords.get("y"))
                                    if rx is not None and ry is not None:
                                        ratio = base_img.width / disp_w
                                        cxd, cyd = x_pos / ratio, y_pos / ratio
                                        changed = False
                                        if "移動" in edit_mode:
                                            nx, ny = int(rx * ratio), int(ry * ratio)
                                            if nx != x_pos or ny != y_pos:
                                                st.session_state[f"px_{slot_id}"] = nx
                                                st.session_state[f"py_{slot_id}"] = ny
                                                changed = True
                                        elif "放大" in edit_mode:
                                            d = math.hypot(rx - cxd, ry - cyd) * ratio
                                            ns = max(10, min(int(d / 2), 200))
                                            if ns != font_size:
                                                st.session_state[f"csz_{slot_id}"] = ns
                                                changed = True
                                        else:
                                            ang = math.degrees(math.atan2(ry - cyd, rx - cxd))
                                            nr = max(-180, min(int(round(ang)), 180))
                                            if nr != rotation_angle:
                                                st.session_state[f"crot_{slot_id}"] = nr
                                                changed = True
                                        if changed:
                                            st.rerun()
                            elif not editor_shown:
                                st.session_state[f"px_{slot_id}"] = x_pos
                                st.session_state[f"py_{slot_id}"] = y_pos
                                st.image(final_img_to_save, width=320)

                            # ================= 📱 手機微調區：完全不用拖曳也能喬位置／大小／角度 =================
                            # 手機上拖曳與拉角縮放很難精準（把手小、又常誤觸捲動），這裡用按鈕一下一下推。
                            # 按下去會同時更新畫布的基準位置並重建畫布，畫布上的文字才會跟著跳到新位置。
                            def _apply_nudge(nx=None, ny=None, nsz=None, nrot=None):
                                _nx = max(0, min(int(x_pos if nx is None else nx), base_img.width))
                                _ny = max(0, min(int(y_pos if ny is None else ny), base_img.height))
                                _ns = max(10, min(int(font_size if nsz is None else nsz), 200))
                                _nr = max(-180, min(int(rotation_angle if nrot is None else nrot), 180))
                                st.session_state[f"px_{slot_id}"] = _nx
                                st.session_state[f"py_{slot_id}"] = _ny
                                st.session_state[f"csz_{slot_id}"] = _ns
                                st.session_state[f"crot_{slot_id}"] = _nr
                                st.session_state[f"cinitx_{slot_id}"] = _nx
                                st.session_state[f"cinity_{slot_id}"] = _ny
                                st.session_state[f"cinitsz_{slot_id}"] = _ns
                                st.session_state[f"cinitrot_{slot_id}"] = _nr
                                st.session_state[f"cvnonce_{slot_id}"] = int(st.session_state.get(f"cvnonce_{slot_id}", 0)) + 1
                                st.rerun()

                            _step_big = st.toggle("粗調（一下移動比較多）", key=f"stepbig_{slot_id}")
                            _mv = max(1, int(base_img.width * (0.05 if _step_big else 0.01)))
                            _dsz = 8 if _step_big else 2
                            _drot = 5 if _step_big else 1

                            _n1, _n2, _n3, _n4 = st.columns(4)
                            _n1.markdown('<span class="keep-row nudge-row" style="display:none;"></span>', unsafe_allow_html=True)
                            if _n1.button("⬅", key=f"ndl_{slot_id}", use_container_width=True, help="往左移"):
                                _apply_nudge(nx=x_pos - _mv)
                            if _n2.button("➡", key=f"ndr_{slot_id}", use_container_width=True, help="往右移"):
                                _apply_nudge(nx=x_pos + _mv)
                            if _n3.button("⬆", key=f"ndu_{slot_id}", use_container_width=True, help="往上移"):
                                _apply_nudge(ny=y_pos - _mv)
                            if _n4.button("⬇", key=f"ndd_{slot_id}", use_container_width=True, help="往下移"):
                                _apply_nudge(ny=y_pos + _mv)

                            _m1, _m2, _m3, _m4 = st.columns(4)
                            _m1.markdown('<span class="keep-row nudge-row" style="display:none;"></span>', unsafe_allow_html=True)
                            if _m1.button("➖字", key=f"ndsm_{slot_id}", use_container_width=True, help="字變小"):
                                _apply_nudge(nsz=font_size - _dsz)
                            if _m2.button("➕字", key=f"ndbg_{slot_id}", use_container_width=True, help="字變大"):
                                _apply_nudge(nsz=font_size + _dsz)
                            if _m3.button("↺", key=f"ndrl_{slot_id}", use_container_width=True, help="逆時針轉"):
                                _apply_nudge(nrot=rotation_angle - _drot)
                            if _m4.button("↻", key=f"ndrr_{slot_id}", use_container_width=True, help="順時針轉"):
                                _apply_nudge(nrot=rotation_angle + _drot)

                            # 📐 大小拉桿：每次重跑都先把拉桿同步成「目前字級」（畫布或 ±字 改過也跟著動），
                            #    並用 on_change 回呼——只有使用者真的拉動才套用。
                            #    （不可改用「比對拉桿值≠字級就套用」：重跑時拉桿可能回到最小值 10，會把字級打回 10。）
                            _szkey = f"szui_{slot_id}"

                            def _on_size_change(_sid=slot_id, _k=_szkey):
                                _v = max(10, min(int(st.session_state.get(_k, 50)), 200))
                                st.session_state[f"csz_{_sid}"] = _v
                                st.session_state[f"cinitsz_{_sid}"] = _v
                                st.session_state[f"cvnonce_{_sid}"] = int(st.session_state.get(f"cvnonce_{_sid}", 0)) + 1

                            st.session_state[_szkey] = int(font_size)
                            # label 只留兩個字：全站 CSS 把拉桿 label 鎖成 60px 並排，長 label 會疊在數值上
                            st.slider("大小", 10, 200, key=_szkey, on_change=_on_size_change)
                            _cent_c1, _cent_c2 = st.columns(2)
                            _cent_c1.markdown('<span class="keep-row nudge-row" style="display:none;"></span>', unsafe_allow_html=True)
                            _do_center = _cent_c1.button("🎯 水平置中", key=f"ndcx_{slot_id}", use_container_width=True, help="文字移到圖片正中間（左右）")
                            _do_reset = _cent_c2.button("⟲ 角度歸零", key=f"ndr0_{slot_id}", use_container_width=True, help="轉正")
                            st.caption(f"目前：位置 {x_pos},{y_pos}　大小 {font_size}　角度 {rotation_angle}°"
                                       f"　·　按一下移動 {_mv}px（{'粗調' if _step_big else '細調'}）")
                            if _do_center:
                                _apply_nudge(nx=base_img.width // 2)
                            if _do_reset:
                                _apply_nudge(nrot=0)
                        else:
                            st.markdown("**👇 原始圖片預覽:**")
                            st.image(base_img, width=300)
                            
                        # 編輯文字 / 完成編輯 ＋ 確認儲存：縮窄併排同一排（僅畫布模式才有編輯鈕）
                        if enable_text and base_img and use_canvas:
                            _bc1, _bc2 = st.columns(2)
                            _bc1.markdown('<span class="keep-row" style="display:none"></span>', unsafe_allow_html=True)
                            if st.session_state.get("edit_slot") == slot_id:
                                if _bc1.button("↩ 完成編輯", key=f"endcv_{slot_id}", use_container_width=True):
                                    st.session_state.edit_slot = None
                                    # 清掉本次編輯的疊字清理暫存，下次編輯從已存位置乾淨起步
                                    for _k in ("cinitx", "cinity", "cinitsz", "cinitrot", "cvnonce"):
                                        st.session_state.pop(f"{_k}_{slot_id}", None)
                                    st.rerun()
                            else:
                                if _bc1.button("🎨 編輯文字", key=f"startcv_{slot_id}", use_container_width=True):
                                    st.session_state.edit_slot = slot_id
                                    for _k in ("cinitx", "cinity", "cinitsz", "cinitrot", "cvnonce"):
                                        st.session_state.pop(f"{_k}_{slot_id}", None)
                                    st.rerun()
                            save_clicked = _bc2.button("✅ 確認儲存", type="primary", use_container_width=True, key=f"btn_save_{slot_id}")
                        else:
                            save_clicked = st.button("✅ 確認儲存", type="primary", use_container_width=True, key=f"btn_save_{slot_id}")
                        if save_clicked:
                            with st.spinner("高畫質切塊處理中，並同步至雲端..."):
                                chunks = img_to_base64_chunks(final_img_to_save)
                                row_data = [f"coupon_{slot_id}"] + chunks
                                
                                rows_to_del = [i + 1 for i, r in enumerate(cfg_data) if r and r[0] == f"coupon_{slot_id}"]
                                if rows_to_del:
                                    for row_index in sorted(rows_to_del, reverse=True):
                                        ws_cfg.delete_rows(row_index)
                                ws_cfg.append_row(row_data)
                                # 🧼 安全網：若還沒有乾淨底圖、且目前底圖是乾淨的(剛上傳)，補存一份乾淨底圖
                                #    這樣即使「上傳後直接加字、沒先按存原圖」，下次編輯也不會疊字。
                                if clean_row is None and new_file is not None:
                                    try:
                                        base_chunks = img_to_base64_chunks(base_img.convert("RGB"))
                                        ws_cfg.append_row([f"couponbase_{slot_id}"] + base_chunks)
                                    except Exception:
                                        pass
                                # 📅 功能3：一併記住壓印位置（下次換日期免重喬）
                                if enable_text:
                                    # 末欄加存「壓印文字」：高畫質存圖才能用大字級重畫同一段字（| 會被當分隔符，換成全形）
                                    _txt_mem = (text_input or "").replace("|", "｜")
                                    _save_kv(ws_cfg, f"coupset_{slot_id}",
                                             f"{x_pos}|{y_pos}|{font_size}|{rotation_angle}|{text_color}|{_txt_mem}")

                                st.session_state.pop("_decoded_imgs", None)
                                st.session_state.refresh_cfg = True
                                st.success("更新成功！")
                                st.rerun()

                st.markdown("<div style='height: 15px;'></div>", unsafe_allow_html=True)

else:
    st.warning("⚠️ 系統尚未連線，請檢查 Secrets 中的金鑰設定。")